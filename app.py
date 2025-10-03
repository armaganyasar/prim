import pandas as pd
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from sqlalchemy import create_engine, text
from config import MYSQL_HOST, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DB, USERS
from dateutil import parser
import logging, pprint
from datetime import datetime, date
from functools import wraps
from database import prim_db, cari_db, personel_db
from prim_utils import *
import json
import io  # Excel için gerekli

app = Flask(__name__)
app.secret_key = "supersecretkey"
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_database_connection():
    return create_engine(
        f"mysql+mysqlconnector://{MYSQL_USER}:{MYSQL_PASSWORD}@{MYSQL_HOST}/{MYSQL_DB}"
    )

# --- Güvenli parametreli sorgu yardımcı fonksiyonu ---
def execute_query(query, params=None):
    """Güvenli parametreli sorgu çalıştırma"""
    engine = get_database_connection()
    try:
        if params:
            result = pd.read_sql(text(query), engine, params=params)
        else:
            result = pd.read_sql(text(query), engine)
        return result
    except Exception as e:
        logger.error(f"Database query error: {e}")
        raise
    finally:
        engine.dispose()

# --- Input validation fonksiyonları ---
def validate_required_fields(data, required_fields):
    """Gerekli alanları kontrol et"""
    for field in required_fields:
        if not data.get(field) or not str(data[field]).strip():
            return False, f"'{field}' alanı gereklidir"
    return True, None

def validate_user_role(role):
    """Kullanıcı rolü doğrulama"""
    valid_roles = ['admin', 'doktor', 'user']
    return role in valid_roles

# --- Login Kontrolü ---
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            flash("Lütfen giriş yapın.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in") or session.get("role") != "admin":
            return jsonify({"error": "Yetkisiz erişim"}), 403
        return f(*args, **kwargs)
    return wrapper

# --- Login ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Kullanıcı adı ve şifre gereklidir", "danger")
            return render_template("login.html")

        user = USERS.get(username)

        if user and user["password"] == password:
            session["logged_in"] = True
            session["username"] = username
            session["role"] = user["role"]
            
            if user["role"] == "doktor":
                session["doktor_id"] = user.get("doktor_id")
            elif user["role"] == "user":
                session["hekimler"] = user.get("hekimler", [])
                
            flash("Giriş başarılı", "success")

            # Yönlendirme: Admin dışındaki herkes randevu sayfasına
            if user["role"] == "admin":
                return redirect(url_for("home"))
            else:
                return redirect(url_for("randevu"))

        flash("Hatalı kullanıcı adı veya şifre", "danger")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Çıkış yapıldı.", "info")
    return redirect(url_for("login"))

@app.route("/")
@login_required
def home():
    if session.get("role") != "admin":
        return redirect(url_for("randevu"))
    return render_template("index.html")

# --- İyileştirilmiş Şubeler API ---
@app.route("/api/branches")
@login_required
def get_branches():
    try:
        query = "SELECT CARI_ID AS id, UNVANI AS name FROM subeler WHERE SILINDI = :silindi ORDER BY UNVANI"
        df = execute_query(query, {"silindi": "false"})
        return jsonify(df.to_dict("records"))
    except Exception as e:
        logger.error(f"Şubeler API hatası: {e}")
        return jsonify({"error": "Şubeler yüklenemedi"}), 500

@app.route("/api/doctors")
@login_required
def get_doctors():
    try:
        sube_id = request.args.get("sube_id")
        
        if sube_id:
            query = """
                SELECT CARI_ID AS id, CONCAT(ADI, ' ', SOYADI) AS name, SUBE_ID, 
                       IFNULL(PRIMYUZDE, 0) AS PRIMYUZDE
                FROM kartdoktor
                WHERE SILINDI = :silindi AND SUBE_ID = :sube_id
                ORDER BY ADI, SOYADI
            """
            params = {"silindi": "false", "sube_id": sube_id}
        else:
            query = """
                SELECT CARI_ID AS id, CONCAT(ADI, ' ', SOYADI) AS name, SUBE_ID, 
                       IFNULL(PRIMYUZDE, 0) AS PRIMYUZDE
                FROM kartdoktor
                WHERE SILINDI = :silindi
                ORDER BY ADI, SOYADI
            """
            params = {"silindi": "false"}
        
        df = execute_query(query, params)
        
        # NaN değerlerini 0 ile değiştir
        df['PRIMYUZDE'] = df['PRIMYUZDE'].fillna(0)
        
        return jsonify(df.to_dict("records"))
    except Exception as e:
        logger.error(f"Doktorlar API hatası: {e}")
        return jsonify({"error": "Doktorlar yüklenemedi"}), 500

# --- Kullanıcıya atanmış hekimler ---
@app.route("/api/me/assignments")
@login_required
def my_assignments():
    if session.get("role") == "user":
        return jsonify(session.get("hekimler", []))
    return jsonify([])

# --- İyileştirilmiş Randevu API ---
@app.route("/api/events")
@login_required
def get_events():
    try:
        start_param = request.args.get("start")
        end_param = request.args.get("end")
        
        if not start_param or not end_param:
            return jsonify([])

        try:
            start_date = parser.parse(start_param).date()
            end_date = parser.parse(end_param).date()
        except ValueError:
            return jsonify({"error": "Geçersiz tarih formatı"}), 400

        # Base query parametreleri
        params = {
            "silindi": "false",
            "start_date": start_date,
            "end_date": end_date
        }

        base_query = """
        SELECT 
            r.ROWNO,
            r.ADISOYADI AS hasta_adi,
            CONCAT(d.ADI, ' ', d.SOYADI) AS doktor_adi,
            s.UNVANI AS sube_adi,
            r.TARIH,
            r.SAATSTART,
            r.DAKKA,
            r.SUBE_ID,
            REPLACE(r.KAYNAKID, 'DOK-', '') AS doktor_id
        FROM randevu r
        LEFT JOIN kartdoktor d ON REPLACE(r.KAYNAKID, 'DOK-', '') = d.CARI_ID
        LEFT JOIN subeler s ON r.SUBE_ID = s.CARI_ID
        WHERE r.SILINDI = :silindi
          AND r.TARIH >= :start_date
          AND r.TARIH < :end_date
        """

        # Rol bazlı filtreleme
        if session.get("role") == "doktor":
            base_query += " AND REPLACE(r.KAYNAKID, 'DOK-', '') = :doktor_id"
            params["doktor_id"] = session.get("doktor_id")
            
        elif session.get("role") == "user":
            sube_id = request.args.get("sube_id")
            doktor_ids_param = request.args.get("doktor_id")
            hekimler = session.get("hekimler", [])

            if sube_id and doktor_ids_param:
                # Kullanıcı filtreleme yaptıysa
                doktor_ids_list = [did.strip() for did in doktor_ids_param.split(',') if did.strip()]
                if doktor_ids_list:
                    placeholders = ','.join([f':doktor_{i}' for i in range(len(doktor_ids_list))])
                    base_query += f" AND r.SUBE_ID = :sube_id AND REPLACE(r.KAYNAKID, 'DOK-', '') IN ({placeholders})"
                    params["sube_id"] = sube_id
                    for i, did in enumerate(doktor_ids_list):
                        params[f"doktor_{i}"] = did
            else:
                # Kullanıcının atamalarına göre
                if hekimler:
                    conditions = []
                    for i, h in enumerate(hekimler):
                        condition = f"(r.SUBE_ID = :user_sube_{i} AND REPLACE(r.KAYNAKID, 'DOK-', '') = :user_doktor_{i})"
                        conditions.append(condition)
                        params[f"user_sube_{i}"] = h['sube_id']
                        params[f"user_doktor_{i}"] = h['doktor_id']
                    
                    if conditions:
                        base_query += " AND (" + " OR ".join(conditions) + ")"

        else:  # Admin
            sube_id = request.args.get("sube_id")
            doktor_ids_param = request.args.get("doktor_id")
            
            if sube_id:
                base_query += " AND r.SUBE_ID = :admin_sube_id"
                params["admin_sube_id"] = sube_id
                
            if doktor_ids_param:
                doktor_ids_list = [did.strip() for did in doktor_ids_param.split(',') if did.strip()]
                if doktor_ids_list:
                    placeholders = ','.join([f':admin_doktor_{i}' for i in range(len(doktor_ids_list))])
                    base_query += f" AND REPLACE(r.KAYNAKID, 'DOK-', '') IN ({placeholders})"
                    for i, did in enumerate(doktor_ids_list):
                        params[f"admin_doktor_{i}"] = did

        df = execute_query(base_query, params)

        if df.empty:
            return jsonify([])

        # Tarih ve saat işleme
        df["start"] = pd.to_datetime(
            df["TARIH"].astype(str).str.strip() + " " + df["SAATSTART"].astype(str).str.strip(),
            format="%Y-%m-%d %H:%M",
            errors="coerce"
        )
        
        df["DAKKA"] = pd.to_numeric(df["DAKKA"], errors="coerce").fillna(30)
        df["end"] = df["start"] + pd.to_timedelta(df["DAKKA"], unit="m")

        # FullCalendar formatına çevir
        events = df.copy()
        events["id"] = events["ROWNO"]
        events["title"] = events["hasta_adi"]
        events["start"] = events["start"].dt.strftime("%Y-%m-%dT%H:%M:%S")
        events["end"] = events["end"].dt.strftime("%Y-%m-%dT%H:%M:%S")
        events["extendedProps"] = events.apply(
            lambda x: {
                "doktor": x["doktor_adi"] or "Belirtilmemiş",
                "sube": x["sube_adi"] or "Belirtilmemiş"
            }, axis=1
        )

        logger.info(f"API çıktısı kayıt sayısı: {len(events)}")
        return jsonify(events[["id", "title", "start", "end", "extendedProps"]].to_dict("records"))

    except Exception as e:
        logger.error(f"Randevu API hatası: {e}")
        return jsonify({"error": "Randevular yüklenemedi"}), 500

# --- Admin kullanıcı listesi ---
@app.route("/api/admin/users")
@admin_required
def admin_users():
    try:
        users_data = []
        for username, info in USERS.items():
            user_info = {
                "username": username,
                "role": info["role"]
            }
            if info["role"] == "doktor":
                user_info["doktor_id"] = info.get("doktor_id")
            elif info["role"] == "user":
                user_info["hekimler"] = info.get("hekimler", [])
            users_data.append(user_info)
        return jsonify(users_data)
    except Exception as e:
        logger.error(f"Kullanıcı listesi API hatası: {e}")
        return jsonify({"error": "Kullanıcı listesi yüklenemedi"}), 500

# --- İyileştirilmiş kullanıcı ekleme ---
@app.route("/api/admin/users/add", methods=["POST"])
@admin_required
def add_user():
    try:
        data = request.get_json() or {}
        
        # Temel alan kontrolü
        valid, error_msg = validate_required_fields(data, ["username", "password", "role"])
        if not valid:
            return jsonify({"error": error_msg}), 400

        username = data["username"].strip()
        password = data["password"].strip()
        role = data["role"].strip()

        # Rol doğrulama
        if not validate_user_role(role):
            return jsonify({"error": "Geçersiz rol"}), 400

        # Kullanıcı varlık kontrolü
        if username in USERS:
            return jsonify({"error": "Bu kullanıcı zaten mevcut"}), 400

        # Temel kullanıcı oluştur
        new_user = {"password": password, "role": role}

        # Rol bazlı özel alanlar
        if role == "doktor":
            doktor_id = data.get("doktor_id", "").strip()
            if not doktor_id:
                return jsonify({"error": "Doktor rolü için 'doktor_id' gerekli"}), 400
            new_user["doktor_id"] = doktor_id

        elif role == "user":
            hekimler = data.get("hekimler", [])
            if not isinstance(hekimler, list) or len(hekimler) == 0:
                return jsonify({"error": "Kullanıcı rolü için en az bir hekim ataması gerekli"}), 400
            
            # Hekim atamalarını doğrula
            validated_hekimler = []
            for hekim in hekimler:
                if not isinstance(hekim, dict):
                    continue
                sube_id = str(hekim.get("sube_id", "")).strip()
                doktor_id = str(hekim.get("doktor_id", "")).strip()
                if sube_id and doktor_id:
                    validated_hekimler.append({"sube_id": sube_id, "doktor_id": doktor_id})
            
            if not validated_hekimler:
                return jsonify({"error": "Geçerli hekim ataması bulunamadı"}), 400
            
            new_user["hekimler"] = validated_hekimler

        # Kullanıcıyı ekle
        USERS[username] = new_user

        # config.py'ye kaydet
        success = save_users_to_config()
        if not success:
            # Hata durumunda geri al
            del USERS[username]
            return jsonify({"error": "Kullanıcı kaydedilemedi"}), 500

        return jsonify({"success": True, "message": "Kullanıcı başarıyla eklendi"})

    except Exception as e:
        logger.error(f"Kullanıcı ekleme hatası: {e}")
        return jsonify({"error": "Kullanıcı eklenirken bir hata oluştu"}), 500

# --- İyileştirilmiş kullanıcı güncelleme ---
@app.route("/api/admin/users/update", methods=["POST"])
@admin_required
def update_user():
    try:
        data = request.get_json() or {}
        
        username = data.get("username", "").strip()
        if not username:
            return jsonify({"error": "Kullanıcı adı gerekli"}), 400
            
        if username not in USERS:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404

        # Şifre güncellemesi
        new_password = data.get("password", "").strip()
        if new_password:
            USERS[username]["password"] = new_password
        
        # Rol güncellemesi
        new_role = data.get("role", "").strip()
        if new_role:
            if not validate_user_role(new_role):
                return jsonify({"error": "Geçersiz rol"}), 400
                
            old_role = USERS[username]["role"]
            USERS[username]["role"] = new_role
            
            # Rol değişmişse eski atamalarını temizle
            if new_role != old_role:
                USERS[username].pop("doktor_id", None)
                USERS[username].pop("hekimler", None)
        
        # Yeni rol bazlı atamalar
        role = USERS[username]["role"]
        
        if role == "doktor":
            doktor_id = data.get("doktor_id", "").strip()
            if not doktor_id:
                return jsonify({"error": "Doktor rolü için 'doktor_id' gerekli"}), 400
            USERS[username]["doktor_id"] = doktor_id
            
        elif role == "user":
            hekimler = data.get("hekimler", [])
            if not isinstance(hekimler, list) or len(hekimler) == 0:
                return jsonify({"error": "Kullanıcı rolü için en az bir hekim ataması gerekli"}), 400
            
            validated_hekimler = []
            for hekim in hekimler:
                if not isinstance(hekim, dict):
                    continue
                sube_id = str(hekim.get("sube_id", "")).strip()
                doktor_id = str(hekim.get("doktor_id", "")).strip()
                if sube_id and doktor_id:
                    validated_hekimler.append({"sube_id": sube_id, "doktor_id": doktor_id})
            
            if not validated_hekimler:
                return jsonify({"error": "Geçerli hekim ataması bulunamadı"}), 400
                
            USERS[username]["hekimler"] = validated_hekimler

        # Kaydet
        success = save_users_to_config()
        if not success:
            return jsonify({"error": "Değişiklikler kaydedilemedi"}), 500

        return jsonify({"success": True, "message": "Kullanıcı başarıyla güncellendi"})

    except Exception as e:
        logger.error(f"Kullanıcı güncelleme hatası: {e}")
        return jsonify({"error": "Kullanıcı güncellenirken bir hata oluştu"}), 500

# --- İyileştirilmiş kullanıcı silme ---
@app.route("/api/admin/users/delete", methods=["POST"])
@admin_required
def delete_user():
    try:
        data = request.get_json() or {}
        username = data.get("username", "").strip()
        
        if not username:
            return jsonify({"error": "Kullanıcı adı gerekli"}), 400
            
        if username not in USERS:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404

        # Kullanıcıyı sil
        del USERS[username]

        # Kaydet
        success = save_users_to_config()
        if not success:
            return jsonify({"error": "Silme işlemi kaydedilemedi"}), 500

        return jsonify({"success": True, "message": "Kullanıcı başarıyla silindi"})

    except Exception as e:
        logger.error(f"Kullanıcı silme hatası: {e}")
        return jsonify({"error": "Kullanıcı silinirken bir hata oluştu"}), 500

def save_users_to_config():
    """USERS dictionary'sini config.py'ye kaydet"""
    try:
        with open("config.py", "r", encoding="utf-8") as f:
            lines = f.readlines()

        # USERS bloğunu bul
        start_idx = end_idx = None
        brace_count = 0
        
        for i, line in enumerate(lines):
            if line.strip().startswith("USERS = {"):
                start_idx = i
                brace_count = line.count("{") - line.count("}")
                continue
            if start_idx is not None:
                brace_count += line.count("{") - line.count("}")
                if brace_count == 0:
                    end_idx = i
                    break

        # Yeni USERS bloğu oluştur
        users_str = pprint.pformat(USERS, indent=4, width=120, sort_dicts=False)
        new_block = f"USERS = {users_str}\n"

        if start_idx is not None and end_idx is not None:
            lines[start_idx:end_idx+1] = [new_block]
        else:
            # USERS bloğu bulunamazsa sona ekle
            lines.append("\n" + new_block)

        # Dosyaya yaz
        with open("config.py", "w", encoding="utf-8") as fw:
            fw.writelines(lines)

        return True
    except Exception as e:
        logger.error(f"Config kaydetme hatası: {e}")
        return False

# --- Randevu HTML route ---
@app.route("/randevu")
@login_required
def randevu():
    return render_template("randevu.html")

# ====================
# PRİM YÖNETİMİ ROUTES
# ====================

@app.route("/primler")
@login_required
def primler():
    """Prim yönetimi ana sayfası"""
    try:
        # Şubeleri getir
        query = "SELECT CARI_ID AS id, UNVANI AS name FROM subeler WHERE SILINDI = :silindi ORDER BY UNVANI"
        df = execute_query(query, {"silindi": "false"})
        branches = df.to_dict("records")
        
        # Ayarları getir
        ayarlar = prim_db.ayarlar_getir()
        
        return render_template("primler.html", 
                             branches=branches, 
                             ayarlar=ayarlar)
        
    except Exception as e:
        logger.error(f"Primler sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu.", "danger")
        return redirect(url_for("home"))

@app.route("/api/prim/tahsilat_getir", methods=["POST"])
@login_required
def prim_tahsilat_getir():
    """Hekim tahsilat verilerini getir"""
    try:
        data = request.get_json() or {}
        
        # Gerekli alanları kontrol et
        valid, error_msg = validate_required_fields(data, ["doktor_id", "baslangic_tarihi", "bitis_tarihi"])
        if not valid:
            return jsonify({"error": error_msg}), 400
        
        mysql_config = {
            'host': MYSQL_HOST,
            'user': MYSQL_USER,
            'password': MYSQL_PASSWORD,
            'db': MYSQL_DB
        }
        
        # Tahsilat verilerini getir
        tahsilat_verileri = get_hekim_tahsilat_verileri(
            data["doktor_id"],
            data["baslangic_tarihi"],
            data["bitis_tarihi"],
            mysql_config
        )
        
        if not tahsilat_verileri:
            return jsonify({"error": "Bu kriterlere uygun tahsilat bulunamadı"}), 404
        
        # Her tahsilat için ödeme şekli analizi
        for tahsilat in tahsilat_verileri:
            tahsilat['odeme_analizi'] = odeme_sekli_analiz(tahsilat['ODEME_SEKLI'])
        
        return jsonify({
            "success": True,
            "data": tahsilat_verileri,
            "doktor_bilgisi": {
                "doktor_id": data["doktor_id"],
                "doktor_adi": tahsilat_verileri[0]['HEKIM_ADI'] if tahsilat_verileri else "",
                "sube_id": tahsilat_verileri[0]['SUBE_ID'] if tahsilat_verileri else "",
                "sube_adi": tahsilat_verileri[0]['SUBE_ADI'] if tahsilat_verileri else "",
                "prim_orani": tahsilat_verileri[0]['PRIMYUZDE'] if tahsilat_verileri else 0
            }
        })
        
    except Exception as e:
        logger.error(f"Tahsilat getirme hatası: {e}")
        return jsonify({"error": "Tahsilat verileri alınırken hata oluştu"}), 500

@app.route("/api/prim/hesapla", methods=["POST"])
@login_required
def prim_hesapla_api():
    """Prim hesaplama API'si - GÜNCELLENMİŞ VERSİYON"""
    try:
        data = request.get_json() or {}
        
        tahsilat_listesi = data.get('tahsilat_listesi', [])
        giderler_listesi = data.get('giderler_listesi', [])
        prim_orani = float(data.get('prim_orani', 0))
        
        # Her tahsilat için kesinti hesapla
        processed_tahsilat = []
        for tahsilat in tahsilat_listesi:
            kesintiler = kesinti_hesapla(
                tahsilat['odeme_sekli'],
                float(tahsilat['brut_tutar']),
                int(tahsilat.get('taksit_sayisi', 1)),
                float(tahsilat.get('kdv_orani', 0)),
                float(tahsilat.get('taksit_kesinti_orani', 0)),
                tahsilat.get('fatura_kesildi', False),
                float(tahsilat.get('pos_komisyon_orani', 0))  # YENİ ALAN
            )
            
            tahsilat_processed = tahsilat.copy()
            tahsilat_processed.update(kesintiler)
            processed_tahsilat.append(tahsilat_processed)
        
        # Prim hesaplama
        hesaplama_sonuc = prim_hesapla(processed_tahsilat, giderler_listesi, prim_orani)
        
        return jsonify({
            "success": True,
            "hesaplama": hesaplama_sonuc,
            "tahsilat_detaylari": processed_tahsilat
        })
        
    except Exception as e:
        logger.error(f"Prim hesaplama hatası: {e}")
        return jsonify({"error": "Prim hesaplanırken hata oluştu"}), 500


@app.route("/api/prim/liste")
@login_required
def prim_liste_api():
    """Kaydedilmiş primleri listele"""
    try:
        doktor_id = request.args.get("doktor_id")
        sube_id = request.args.get("sube_id")  # EKLE
        baslangic = request.args.get("baslangic")
        bitis = request.args.get("bitis")
        page = int(request.args.get("page", 1))
        per_page = 20
        
        # SQLite veritabanından primleri getir
        primler = prim_db.prim_listele(doktor_id, baslangic, bitis)
        
        # Şube filtrelemesi ekle
        if sube_id:
            primler = [p for p in primler if str(p.get('sube_id')) == str(sube_id)]
        
        # Rol bazlı filtreleme
        if session.get("role") == "doktor":
            session_doktor_id = session.get("doktor_id")
            primler = [p for p in primler if str(p.get('doktor_id')) == str(session_doktor_id)]
        elif session.get("role") == "user":
            hekimler = session.get("hekimler", [])
            allowed_combinations = [(h['sube_id'], h['doktor_id']) for h in hekimler]
            primler = [p for p in primler if (str(p.get('sube_id')), str(p.get('doktor_id'))) in allowed_combinations]
        
        # Sayfalama
        total_records = len(primler)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        
        # Admin silme yetkisi
        is_admin = session.get("role") == "admin"
        for prim in primler:
            prim["can_delete"] = is_admin
        
        return jsonify({
            "success": True,
            "data": primler[start_idx:end_idx],
            "pagination": {
                "current_page": page,
                "total_pages": (total_records + per_page - 1) // per_page,
                "total_records": total_records,
                "per_page": per_page
            },
            "summary": {
                "total_records": total_records,
                "total_amount": sum(float(p.get("hesaplanan_prim", 0)) for p in primler)
            }
        })
        
    except Exception as e:
        logger.error(f"Prim listeleme hatası: {e}")
        return jsonify({"error": "Primler yüklenirken hata oluştu: " + str(e)}), 500
        
@app.route("/api/prim/detay/<int:prim_id>")
@login_required
def prim_detay_api(prim_id):
    """Prim detayını getir"""
    try:
        # Yetki kontrolü - SQLite için basitleştirilmiş
        if session.get("role") != "admin":
            detay = prim_db.prim_detay_getir(prim_id)
            if not detay:
                return jsonify({"error": "Prim kaydı bulunamadı"}), 404
                
            prim_data = detay['prim_data']
            user_role = session.get("role")
            
            if user_role == "doktor":
                if str(prim_data.get('doktor_id')) != str(session.get("doktor_id")):
                    return jsonify({"error": "Bu prime erişim yetkiniz yok"}), 403
            elif user_role == "user":
                hekimler = session.get("hekimler", [])
                allowed = False
                for h in hekimler:
                    if (str(h['sube_id']) == str(prim_data.get('sube_id')) and 
                        str(h['doktor_id']) == str(prim_data.get('doktor_id'))):
                        allowed = True
                        break
                if not allowed:
                    return jsonify({"error": "Bu prime erişim yetkiniz yok"}), 403
        else:
            detay = prim_db.prim_detay_getir(prim_id)
        
        if not detay:
            return jsonify({"error": "Prim kaydı bulunamadı"}), 404
        
        return jsonify({
            "success": True,
            "data": detay
        })
        
    except Exception as e:
        logger.error(f"Prim detay getirme hatası: {e}")
        return jsonify({"error": "Prim detayı yüklenirken hata oluştu"}), 500


@app.route("/api/prim/sil", methods=["POST"])
@admin_required
def prim_sil():
    """Prim kaydını sil - Admin şifre kontrolü ve CARİ HAREKET SİLME ile"""
    try:
        data = request.get_json() or {}
        
        prim_id = data.get("prim_id")
        admin_sifre = data.get("admin_sifre", "").strip()
        
        if not prim_id:
            return jsonify({"error": "Prim ID gerekli"}), 400
            
        if not admin_sifre:
            return jsonify({"error": "Admin şifresi gerekli"}), 400
        
        # Admin şifresini kontrol et
        current_user = session.get("username")
        if current_user not in USERS:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 403
            
        if USERS[current_user]["password"] != admin_sifre:
            return jsonify({"error": "Hatalı admin şifresi"}), 403
        
        # Prim kaydının var olup olmadığını kontrol et (SQLite)
        detay = prim_db.prim_detay_getir(prim_id)
        if not detay:
            return jsonify({"error": "Prim kaydı bulunamadı"}), 404
        
        # Cari Hareketleri ve Prim Kayıtlarını silme işlemi (SQLite)
        try:
            import sqlite3
            conn_prim = sqlite3.connect(prim_db.db_path)
            cursor = conn_prim.cursor()
            
            # 1. Cari Hareketi Silme (İlişkili hareket varsa)
            # prim_id'si ile eşleşen hareketi bul
            cursor.execute("SELECT cari_id FROM cari_hareketler WHERE prim_id = ?", (prim_id,))
            cari_kaydi = cursor.fetchone()

            if cari_kaydi:
                cari_id = cari_kaydi[0]
                
                # Cari hareketi sil
                cursor.execute("DELETE FROM cari_hareketler WHERE prim_id = ?", (prim_id,))
                
                # Prim veritabanındaki değişiklikleri kaydet
                conn_prim.commit() 
                conn_prim.close()

                # Cari Database üzerinden bakiyeyi yeniden hesapla (Farklı bir bağlantı gerektirir)
                success, message = cari_db._recalculate_bakiye(cari_id, sqlite3.connect(cari_db.db_path))
                if not success:
                     logger.error(f"Prim silindi ancak cari bakiye güncellenemedi: Cari ID {cari_id}. Hata: {message}")
                     return jsonify({
                        "success": True, 
                        "message": "Prim başarıyla silindi ancak cari bakiyesi manuel düzeltme gerektirebilir.",
                        "cari_warning": True
                    })
                
                conn_prim = sqlite3.connect(prim_db.db_path) # Yeni bağlantı aç
                cursor = conn_prim.cursor()
            
            # 2. Prim Detaylarını Sil
            cursor.execute("DELETE FROM prim_giderler WHERE prim_id = ?", (prim_id,))
            cursor.execute("DELETE FROM prim_tahsilat_detaylari WHERE prim_id = ?", (prim_id,))
            
            # 3. Ana Prim Kaydını Sil
            cursor.execute("DELETE FROM prim_hesaplamalari WHERE id = ?", (prim_id,))
            
            conn_prim.commit()
            conn_prim.close()
            
            logger.info(f"Prim kaydı ve ilişkili cari hareket silindi: ID={prim_id}, Admin={current_user}")
            return jsonify({
                "success": True,
                "message": "Prim kaydı ve ilişkili cari hareketi başarıyla silindi"
            })
            
        except Exception as delete_error:
            logger.error(f"Silme işlemi hatası: {delete_error}")
            return jsonify({"error": "Prim kaydı ve cari hareket silinirken hata oluştu"}), 500
        
    except Exception as e:
        logger.error(f"Prim silme hatası: {e}")
        return jsonify({"error": "Silme işlemi sırasında hata oluştu"}), 500
        
@app.route("/api/prim/ayarlar")
@login_required
def prim_ayarlar_api():
    """Prim ayarlarını getir"""
    try:
        ayarlar = prim_db.ayarlar_getir()
        return jsonify({
            "success": True,
            "data": ayarlar
        })
        
    except Exception as e:
        logger.error(f"Ayarlar getirme hatası: {e}")
        return jsonify({"error": "Ayarlar yüklenirken hata oluştu"}), 500

@app.route("/api/prim/ayarlar/guncelle", methods=["POST"])
@admin_required
def prim_ayarlar_guncelle():
    """Prim ayarlarını güncelle"""
    try:
        data = request.get_json() or {}
        
        taksit_oranlari = data.get('taksit_oranlari', [])
        gider_kategorileri = data.get('gider_kategorileri', [])
        
        # Veri doğrulama
        if not isinstance(taksit_oranlari, list):
            return jsonify({"error": "Taksit oranları liste formatında olmalıdır"}), 400
            
        if not isinstance(gider_kategorileri, list):
            return jsonify({"error": "Gider kategorileri liste formatında olmalıdır"}), 400
        
        # Taksit oranları doğrulama
        for taksit in taksit_oranlari:
            if not isinstance(taksit.get('taksit_sayisi'), int) or taksit['taksit_sayisi'] < 1:
                return jsonify({"error": "Geçersiz taksit sayısı"}), 400
                
            if not isinstance(taksit.get('kesinti_orani'), (int, float)) or taksit['kesinti_orani'] < 0:
                return jsonify({"error": "Geçersiz kesinti oranı"}), 400
        
        # Gider kategorileri doğrulama
        for kategori in gider_kategorileri:
            if not isinstance(kategori.get('kategori'), str) or not kategori['kategori'].strip():
                return jsonify({"error": "Geçersiz kategori adı"}), 400
        
        # Ayarları güncelle
        success = prim_db.ayarlar_guncelle(taksit_oranlari, gider_kategorileri)
        
        if success:
            return jsonify({"success": True, "message": "Ayarlar başarıyla güncellendi"})
        else:
            return jsonify({"error": "Ayarlar güncellenirken hata oluştu"}), 500
        
    except Exception as e:
        logger.error(f"Ayar güncelleme hatası: {e}")
        return jsonify({"error": "Ayarlar güncellenemedi"}), 500

# --- PRİM LİSTESİ SAYFASI ---
@app.route("/prim_listesi")
@login_required
def prim_listesi():
    """Prim listesi sayfası"""
    try:
        # Şubeleri getir
        query = "SELECT CARI_ID AS id, UNVANI AS name FROM subeler WHERE SILINDI = :silindi ORDER BY UNVANI"
        df = execute_query(query, {"silindi": "false"})
        branches = df.to_dict("records")
        
        return render_template("prim_listesi.html", branches=branches)
        
    except Exception as e:
        logger.error(f"Prim listesi sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu.", "danger")
        return redirect(url_for("home"))

@app.route("/api/prim/ozet_yazdir/<int:prim_id>")
@login_required
def prim_ozet_yazdir(prim_id):
    """Prim özeti yazdırma sayfası"""
    try:
        # Detayları çek
        detay = prim_db.prim_detay_getir(prim_id)
        
        if not detay:
            return "Prim kaydı bulunamadı", 404
            
        # Yetki kontrolü
        if session.get("role") != "admin":
            prim_data = detay['prim_data']
            user_role = session.get("role")
            
            if user_role == "doktor":
                if str(prim_data.get('doktor_id')) != str(session.get("doktor_id")):
                    return "Bu prime erişim yetkiniz yok", 403
            elif user_role == "user":
                hekimler = session.get("hekimler", [])
                allowed = False
                for h in hekimler:
                    if (str(h['sube_id']) == str(prim_data.get('sube_id')) and 
                        str(h['doktor_id']) == str(prim_data.get('doktor_id'))):
                        allowed = True
                        break
                if not allowed:
                    return "Bu prime erişim yetkiniz yok", 403
        
        # Template için veri yapısını DOĞRU şekilde düzenle
        detay_for_template = {
            'prim': detay.get('prim_data'),
            'tahsilat': detay.get('tahsilat_detaylari'),
            'laboratuvar_giderleri': detay.get('laboratuvar_giderleri', []),
            'implant_giderleri': detay.get('implant_giderleri', []),
            'diger_giderler': detay.get('diger_giderler', []),
            'net_ciro_eklemeleri': detay.get('net_ciro_eklemeleri', []),  # YENİ
            'hakedis_eklemeleri': detay.get('hakedis_eklemeleri', [])      # YENİ
        }
        
        logger.info(f"Yazdırma için veri hazırlandı - Prim ID: {prim_id}")
        logger.info(f"Lab gider sayısı: {len(detay_for_template['laboratuvar_giderleri'])}")
        logger.info(f"İmplant gider sayısı: {len(detay_for_template['implant_giderleri'])}")
        logger.info(f"Diğer gider sayısı: {len(detay_for_template['diger_giderler'])}")
        logger.info(f"Net Ciro sayısı: {len(detay_for_template['net_ciro_eklemeleri'])}")  # YENİ
        logger.info(f"Hak Ediş sayısı: {len(detay_for_template['hakedis_eklemeleri'])}")    # YENİ
        
        return render_template("prim_ozet_yazdir.html", detay=detay_for_template)
        
    except Exception as e:
        logger.error(f"Prim özet yazdırma hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return f"Bir hata oluştu: {str(e)}", 500
        # --- TAHSİLATLAR SAYFASİ ---
@app.route("/tahsilatlar")
@login_required
def tahsilatlar():
    try:
        engine = get_database_connection()
        # Şubeleri getir
        branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
        branches_df = pd.read_sql(branches_query, engine)
        branches = branches_df.to_dict('records')
        
        return render_template("tahsilatlar.html", 
                             branches=branches,
                             table_data=None,
                             summary=None,
                             multi_doctor_issues=None,
                             start_date=None,
                             end_date=None,
                             selected_branches=None,
                             selected_doctors=None,
                             error_date=None)
    except Exception as e:
        logger.error(f"Tahsilatlar sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu.", "danger")
        return redirect(url_for("home"))
# Bu kodu app.py dosyanıza ekleyin (prim ile ilgili diğer route'ların yanına)

# Bu kodu app.py dosyanızdaki mevcut check_existing fonksiyonunun yerine koyun

# Bu kodu app.py dosyanızdaki mevcut check_existing fonksiyonunun yerine koyun

@app.route('/api/prim/check_existing', methods=['POST'])
def check_existing_prim():
    """
    Çakışan prim hesaplaması kontrolü
    Aynı hekim için tarih aralığı çakışması var mı kontrol eder
    """
    try:
        data = request.get_json()
        
        doktor_id = data.get('doktor_id')
        baslangic_tarihi = data.get('baslangic_tarihi') 
        bitis_tarihi = data.get('bitis_tarihi')
        
        print(f"Prim çakışma kontrolü: Hekim={doktor_id}, Tarih={baslangic_tarihi} - {bitis_tarihi}")
        
        # Parametreleri kontrol et
        if not all([doktor_id, baslangic_tarihi, bitis_tarihi]):
            return jsonify({
                'success': False,
                'error': 'Eksik parametreler'
            }), 400
        
        # *** DÜZELTİLMİŞ: Doğru veritabanı yolu ve tablo adı ***
        import sqlite3
        import os
        
        # Veritabanı yolunu kontrol et
        db_paths = [
            'data/prim_hesaplamalari.db',  # database.py'deki yol
            'prim_data.db',                # app.py'deki yol
            'prim_hesaplamalari.db'        # alternatif
        ]
        
        db_path = None
        for path in db_paths:
            if os.path.exists(path):
                db_path = path
                print(f"Veritabanı bulundu: {path}")
                break
        
        if not db_path:
            print("Hiçbir veritabanı dosyası bulunamadı")
            return jsonify({
                'success': True,
                'exists': False,
                'count': 0,
                'conflicting_prims': [],
                'message': 'Veritabanı bulunamadı - çakışma kontrolü yapılamadı'
            })
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Önce tabloları listele
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        print(f"Mevcut tablolar: {[table[0] for table in tables]}")
        
        # Doğru tablo adını bul
        table_name = None
        possible_tables = ['prim_hesaplamalari', 'primler', 'prim_kayitlari']
        
        for table in possible_tables:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
            if cursor.fetchone():
                table_name = table
                print(f"Prim tablosu bulundu: {table}")
                break
        
        if not table_name:
            conn.close()
            print("Prim tablosu bulunamadı")
            return jsonify({
                'success': True,
                'exists': False,
                'count': 0,
                'conflicting_prims': [],
                'message': 'Prim tablosu bulunamadı'
            })
        
        # Çakışan primleri bul
        query = f"""
        SELECT 
            id,
            doktor_adi,
            donem_baslangic,
            donem_bitis,
            hesaplanan_prim,
            olusturma_tarihi
        FROM {table_name}
        WHERE doktor_id = ? 
        AND (
            -- Çakışma durumları
            (? BETWEEN donem_baslangic AND donem_bitis) OR  -- Yeni başlangıç mevcut aralıkta
            (? BETWEEN donem_baslangic AND donem_bitis) OR  -- Yeni bitiş mevcut aralıkta
            (donem_baslangic BETWEEN ? AND ?) OR           -- Mevcut başlangıç yeni aralıkta
            (donem_bitis BETWEEN ? AND ?)                  -- Mevcut bitiş yeni aralıkta
        )
        ORDER BY donem_baslangic DESC
        """
        
        cursor.execute(query, (
            doktor_id,
            baslangic_tarihi, bitis_tarihi,  # Yeni tarihler
            baslangic_tarihi, bitis_tarihi,  # Yeni tarihler tekrar
            baslangic_tarihi, bitis_tarihi   # Yeni tarihler tekrar
        ))
        
        conflicting_prims = cursor.fetchall()
        conn.close()
        
        print(f"Sorgu sonucu: {len(conflicting_prims)} çakışan prim bulundu")
        
        # Sonuçları formatla
        if conflicting_prims:
            conflicts = []
            for prim in conflicting_prims:
                conflicts.append({
                    'id': prim[0],
                    'doktor_adi': prim[1],
                    'donem_baslangic': prim[2],
                    'donem_bitis': prim[3],
                    'hesaplanan_prim': float(prim[4]) if prim[4] else 0,
                    'olusturma_tarihi': prim[5]
                })
            
            print(f"Çakışan prim bulundu: {len(conflicts)} adet")
            
            return jsonify({
                'success': True,
                'exists': True,
                'count': len(conflicts),
                'conflicting_prims': conflicts,
                'message': f'{len(conflicts)} adet çakışan prim hesaplaması bulundu'
            })
        else:
            print("Çakışan prim bulunamadı")
            
            return jsonify({
                'success': True,
                'exists': False,
                'count': 0,
                'conflicting_prims': [],
                'message': 'Çakışan prim bulunamadı'
            })
            
    except Exception as e:
        print(f"Prim çakışma kontrol hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Çakışma kontrolü sırasında hata: {str(e)}'
        }), 500
@app.route('/api/prim/conflict_analysis', methods=['POST'])
def prim_conflict_analysis():
    """
    Çakışan primler için detaylı analiz
    Hangi günlerin çakıştığını, toplam tutarları vs. gösterir
    """
    try:
        data = request.get_json()
        
        doktor_id = data.get('doktor_id')
        baslangic_tarihi = data.get('baslangic_tarihi')
        bitis_tarihi = data.get('bitis_tarihi')
        
        # Çakışan primleri bul (yukardaki fonksiyonun aynısı)
        conn = sqlite3.connect('prim_data.db')
        cursor = conn.cursor()
        
        query = """
        SELECT 
            id, doktor_adi, donem_baslangic, donem_bitis, 
            hesaplanan_prim, brut_tahsilat, toplam_gider
        FROM primler 
        WHERE doktor_id = ? 
        AND (
            (? BETWEEN donem_baslangic AND donem_bitis) OR
            (? BETWEEN donem_baslangic AND donem_bitis) OR
            (donem_baslangic BETWEEN ? AND ?) OR
            (donem_bitis BETWEEN ? AND ?)
        )
        """
        
        cursor.execute(query, (
            doktor_id,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi
        ))
        
        conflicting_prims = cursor.fetchall()
        conn.close()
        
        if not conflicting_prims:
            return jsonify({
                'success': True,
                'has_conflicts': False,
                'analysis': 'Çakışma bulunamadı'
            })
        
        # Detaylı analiz
        analysis = {
            'total_conflicting_prims': len(conflicting_prims),
            'total_amount': sum(float(p[4]) for p in conflicting_prims if p[4]),
            'date_range_requested': f"{baslangic_tarihi} - {bitis_tarihi}",
            'conflicts': []
        }
        
        for prim in conflicting_prims:
            # Çakışan günleri hesapla
            existing_start = datetime.strptime(prim[2], '%Y-%m-%d').date()
            existing_end = datetime.strptime(prim[3], '%Y-%m-%d').date()
            new_start = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            new_end = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            
            # Çakışan aralığı bul
            overlap_start = max(existing_start, new_start)
            overlap_end = min(existing_end, new_end)
            overlap_days = (overlap_end - overlap_start).days + 1
            
            analysis['conflicts'].append({
                'prim_id': prim[0],
                'existing_period': f"{prim[2]} - {prim[3]}",
                'overlap_period': f"{overlap_start} - {overlap_end}",
                'overlap_days': overlap_days,
                'amount': float(prim[4]) if prim[4] else 0
            })
        
        return jsonify({
            'success': True,
            'has_conflicts': True,
            'analysis': analysis
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Analiz hatası: {str(e)}'
        }), 500
@app.route("/api/get_doctors")
@login_required  
def get_doctors_by_branches():
    try:
        branch_ids = request.args.get("branch_ids", "").split(",")
        branch_ids = [b.strip() for b in branch_ids if b.strip()]
        
        if not branch_ids:
            return jsonify({"error": "Şube ID'leri gerekli"})
        
        engine = get_database_connection()
        
        # Güvenli parametreli sorgu
        placeholders = ','.join([':branch_' + str(i) for i in range(len(branch_ids))])
        query = f"""
            SELECT CARI_ID, CONCAT(ADI, ' ', SOYADI) AS HEKIM_ADI, SUBE_ID
            FROM kartdoktor 
            WHERE SILINDI='false' AND SUBE_ID IN ({placeholders})
            ORDER BY ADI, SOYADI
        """
        
        params = {}
        for i, bid in enumerate(branch_ids):
            params[f'branch_{i}'] = bid
        
        df = pd.read_sql(text(query), engine, params=params)
        return jsonify(df.to_dict('records'))
        
    except Exception as e:
        logger.error(f"Doktor listesi getirme hatası: {e}")
        return jsonify({"error": "Doktor listesi alınamadı"})

@app.route("/analyze_tahsilatlar", methods=["POST"])
@login_required
def analyze_tahsilatlar():
    try:
        start_date_raw = request.form.get("start_date")
        end_date_raw = request.form.get("end_date")
        selected_branches = request.form.getlist("branches")
        selected_doctors = request.form.getlist("doctors")
        
        if not start_date_raw or not end_date_raw:
            flash("Başlangıç ve bitiş tarihi gereklidir.", "danger")
            return redirect(url_for("tahsilatlar"))

        try:
            start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_raw, "%Y-%m-%d").date()
        except ValueError:
            flash("Tarih formatı hatalı.", "danger")
            return redirect(url_for("tahsilatlar"))

        engine = get_database_connection()

        # Ödeme kesinti oranları
        KESINTI_ORANLARI = {
            'nakit': 0,
            'pos': 0.10,
            'banka': 0.10,
            'çek': 0.10,
            'senet': 0,
            'senet tahsilatı': 0
        }

        # Ana Tahsilat Sorgusu (Güvenli parametreli versiyon)
        branch_placeholders = ','.join([':branch_' + str(i) for i in range(len(selected_branches))])
        
        params = {
            'start_date': start_date,
            'end_date': end_date
        }
        
        # Branch parametrelerini ekle
        for i, branch in enumerate(selected_branches):
            params[f'branch_{i}'] = branch
        
        tahsilat_query = f"""
        SELECT
            T1.TARIH,
            T1.ALACAK AS TUTAR,
            T1.HEDEF_ILGILI_DOKTOR_ID AS DOKTOR_ID,
            CONCAT(IFNULL(DR.ADI,''),' ',IFNULL(DR.SOYADI,'')) AS HEKIM_ADI,
            IFNULL(DR.PRIMYUZDE,0) AS PRIMYUZDE,
            CONCAT(IFNULL(H.ADI,''),' ',IFNULL(H.SOYADI,'')) AS HASTA_ADI,
            H.HASTA_ID,
            LOWER(IFNULL(OS.ADI,'bilinmeyen')) AS ODEME_SEKLI,
            IFNULL(SB.UNVANI,'Bilinmeyen') AS SUBE_ADI,
            T1.HAREKETTYPE
        FROM carihareket AS T1
        LEFT JOIN kartdoktor AS DR ON T1.HEDEF_ILGILI_DOKTOR_ID = DR.CARI_ID
        LEFT JOIN odeme_sekilleri AS OS ON T1.ISLEM_TIPI_ID = OS.ROWNO
        LEFT JOIN subeler AS SB ON T1.SUBE_ID = SB.CARI_ID
        LEFT JOIN karthasta AS H ON T1.KAYNAK_CARI_ID = H.HASTA_ID
        WHERE T1.SILINDI='False'
          AND T1.ALACAK > 0
          AND T1.HAREKETTYPE IN ('T','ST','CT')
          AND T1.TARIH BETWEEN :start_date AND :end_date
          AND T1.SUBE_ID IN ({branch_placeholders})
        """
        
        if selected_doctors:
            doctor_placeholders = ','.join([':doctor_' + str(i) for i in range(len(selected_doctors))])
            tahsilat_query += f" AND T1.HEDEF_ILGILI_DOKTOR_ID IN ({doctor_placeholders})"
            for i, doctor in enumerate(selected_doctors):
                params[f'doctor_{i}'] = doctor

        df = pd.read_sql(text(tahsilat_query), engine, params=params)
        
        if df.empty:
            engine = get_database_connection()
            branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
            branches_df = pd.read_sql(branches_query, engine)
            branches = branches_df.to_dict('records')
            return render_template("tahsilatlar.html", 
                                 branches=branches,
                                 error_date="Bu kriterlere uygun tahsilat bulunamadı.")

        # Kesinti & Prim Hesaplama
        df["ODEME_SEKLI"] = df["ODEME_SEKLI"].str.strip().str.lower()
        df["KESINTI_ORANI"] = df["ODEME_SEKLI"].map(KESINTI_ORANLARI).fillna(0)
        df["KESINTI"] = df["TUTAR"] * df["KESINTI_ORANI"]
        df["NET_TAHSILAT"] = df["TUTAR"] - df["KESINTI"]
        df["PRIM_YUZDE"] = df["PRIMYUZDE"] / 100
        df["HESAPLANAN_PRIM"] = df["NET_TAHSILAT"] * df["PRIM_YUZDE"]

        # ÖNCEKİ TÜM VERİ İLE ÇOKLU HEKİM TESPİTİ YAP (filtre uygulamadan)
        # GÜNCELLENMİŞ TEDAVİ SORGUSU - HEM SILINDI HEM ISDELETED KONTROLLÜ
        tum_tedavi_query = """
        SELECT
            T1.HASTA_ID,
            T1.DOKTOR_ID,
            COALESCE(CONCAT(DR.ADI,' ',DR.SOYADI), '') AS HEKIM_ADI,
            SUM(COALESCE(T1.TUTAR, T1.LISTETUTAR, 0)) AS TEDAVI_TOPLAM,
            COUNT(*) AS TEDAVI_SAYISI,
            MIN(T1.TARIH) AS ILK_TEDAVI,
            MAX(T1.TARIH) AS SON_TEDAVI,
            DR.SUBE_ID
        FROM tedavi AS T1
        LEFT JOIN kartdoktor AS DR ON T1.DOKTOR_ID = DR.CARI_ID
        WHERE (T1.SILINDI IS NULL OR T1.SILINDI != 'True')
          AND (T1.ISDELETED IS NULL OR T1.ISDELETED != 'True')
        GROUP BY T1.HASTA_ID, T1.DOKTOR_ID, HEKIM_ADI, DR.SUBE_ID
        """
        tum_tedavi_df = pd.read_sql(tum_tedavi_query, engine)

        # TÜM VERİDEN ÇOKLU HEKİM HASTALARINI BUL
        tum_multi_patients = set(
            tum_tedavi_df.groupby("HASTA_ID")["DOKTOR_ID"].nunique()
            .loc[lambda x: x > 1].index.astype(str).tolist()
        )

        # FİLTRE 1: Ana analizde görünen hastalarla sınırla
        ana_analiz_hastalari = set(df["HASTA_ID"].astype(str).tolist())
        multi_patients = tum_multi_patients.intersection(ana_analiz_hastalari)
        
        # FİLTRE 2: Seçilen hekimlerle ilgili çoklu hekim durumlarını bul
        if selected_doctors:
            # Seçilen hekimlerden en az birinden tedavi gören çoklu hekim hastalarını al
            selected_doctor_patients = set(
                tum_tedavi_df[tum_tedavi_df["DOKTOR_ID"].isin(selected_doctors)]["HASTA_ID"].astype(str).tolist()
            )
            multi_patients = multi_patients.intersection(selected_doctor_patients)

        # FİLTRE 3: Seçilen şubelerle ilgili çoklu hekim durumlarını bul
        selected_branch_patients = set(
            tum_tedavi_df[tum_tedavi_df["SUBE_ID"].isin(selected_branches)]["HASTA_ID"].astype(str).tolist()
        )
        multi_patients = multi_patients.intersection(selected_branch_patients)

        logger.info(f"Toplam çoklu hekim hasta sayısı: {len(tum_multi_patients)}")
        logger.info(f"Ana analizde bulunan çoklu hekim hasta sayısı: {len(tum_multi_patients.intersection(ana_analiz_hastalari))}")
        logger.info(f"Filtreler uygulandıktan sonra: {len(multi_patients)}")

        # PERFORMANS İYİLEŞTİRMESİ: Eğer çoklu hekim hasta sayısı çok fazlaysa sınırla
        if len(multi_patients) > 50:
            logger.warning(f"Çok fazla çoklu hekim hastası bulundu ({len(multi_patients)}). İlk 50 tanesi gösterilecek.")
            multi_patients = set(list(multi_patients)[:50])

        # Ana analiz için sadece ilgili verileri al
        tedavi_df = tum_tedavi_df[tum_tedavi_df["HASTA_ID"].astype(str).isin(multi_patients)]

        # TÜM ZAMANLARI KAPSAYAN TAHSİLAT SORGUSU (Çoklu hekim hastalarına odaklı)
        if multi_patients:
            hasta_ids_list = list(multi_patients)
            hasta_placeholders = ','.join([':hasta_' + str(i) for i in range(len(hasta_ids_list))])
            
            tum_tahsilat_params = {}
            for i, pid in enumerate(hasta_ids_list):
                tum_tahsilat_params[f'hasta_{i}'] = pid
                
            tum_tahsilat_query = f"""
            SELECT
                T1.HEDEF_ILGILI_DOKTOR_ID AS DOKTOR_ID,
                H.HASTA_ID,
                SUM(T1.ALACAK) AS TOPLAM_TUTAR,
                COUNT(*) AS TAHSILAT_SAYISI,
                MAX(T1.TARIH) AS SON_TAHSILAT_TARIHI
            FROM carihareket AS T1
            LEFT JOIN karthasta AS H ON T1.KAYNAK_CARI_ID = H.HASTA_ID
            WHERE T1.SILINDI='False'
              AND T1.ALACAK > 0
              AND T1.HAREKETTYPE IN ('T','ST','CT')
              AND H.HASTA_ID IN ({hasta_placeholders})
            GROUP BY T1.HEDEF_ILGILI_DOKTOR_ID, H.HASTA_ID
            """
            tum_tahsilat_df = pd.read_sql(text(tum_tahsilat_query), engine, params=tum_tahsilat_params)
        else:
            tum_tahsilat_df = pd.DataFrame()

        # Çoklu hekim işareti ekle (sadece filtrelenmiş veride)
        df["IS_MULTI_DOCTOR"] = df["HASTA_ID"].astype(str).isin(multi_patients)

        # PERFORMANS İYİLEŞTİRMESİ: Hasta ve hekim adlarını toplu olarak çek
        hasta_adlari_dict = {}
        hekim_adlari_dict = {}
        
        if multi_patients:
            # Tüm hasta adlarını tek sorguda çek
            hasta_ids_list = list(multi_patients)
            hasta_placeholders = ','.join([':hasta_name_' + str(i) for i in range(len(hasta_ids_list))])
            hasta_name_params = {}
            for i, pid in enumerate(hasta_ids_list):
                hasta_name_params[f'hasta_name_{i}'] = pid
                
            tum_hasta_adlari_query = f"""
            SELECT HASTA_ID, CONCAT(IFNULL(ADI,''), ' ', IFNULL(SOYADI,'')) AS HASTA_ADI 
            FROM karthasta 
            WHERE HASTA_ID IN ({hasta_placeholders})
            """
            hasta_adlari_df = pd.read_sql(text(tum_hasta_adlari_query), engine, params=hasta_name_params)
            hasta_adlari_dict = dict(zip(hasta_adlari_df["HASTA_ID"].astype(str), hasta_adlari_df["HASTA_ADI"]))

            # Sadece çoklu hekim hastalarının hekim adlarını çek
            filtered_tahsilat = tum_tahsilat_df[tum_tahsilat_df["HASTA_ID"].astype(str).isin(multi_patients)]
            filtered_tedavi = tedavi_df[tedavi_df["HASTA_ID"].astype(str).isin(multi_patients)]
            
            tahsilat_hekim_ids = set(filtered_tahsilat["DOKTOR_ID"].tolist())
            tedavi_hekim_ids = set(filtered_tedavi["DOKTOR_ID"].tolist())
            tum_hekim_ids = tahsilat_hekim_ids.union(tedavi_hekim_ids)
            
            if tum_hekim_ids:
                hekim_ids_list = list(tum_hekim_ids)
                hekim_placeholders = ','.join([':hekim_' + str(i) for i in range(len(hekim_ids_list))])
                hekim_params = {}
                for i, hid in enumerate(hekim_ids_list):
                    hekim_params[f'hekim_{i}'] = hid
                    
                tum_hekim_adlari_query = f"""
                SELECT CARI_ID, CONCAT(IFNULL(ADI,''), ' ', IFNULL(SOYADI,'')) AS HEKIM_ADI 
                FROM kartdoktor 
                WHERE CARI_ID IN ({hekim_placeholders})
                """
                hekim_adlari_df = pd.read_sql(text(tum_hekim_adlari_query), engine, params=hekim_params)
                hekim_adlari_dict = dict(zip(hekim_adlari_df["CARI_ID"], hekim_adlari_df["HEKIM_ADI"]))

        # Çoklu hekim detayları (Optimize edilmiş versiyon)
        multi_doctor_issues = []
        if multi_patients:
            for hasta_id in multi_patients:
                hasta_tedavi = tedavi_df[tedavi_df["HASTA_ID"].astype(str) == str(hasta_id)]
                hasta_tahsilat = df[df["HASTA_ID"].astype(str) == str(hasta_id)]
                
                # Hasta adını bul (önce df'den, sonra dict'ten)
                if not hasta_tahsilat.empty:
                    hasta_adi = hasta_tahsilat.iloc[0]["HASTA_ADI"]
                else:
                    hasta_adi = hasta_adlari_dict.get(str(hasta_id), f"Hasta ID: {hasta_id}")

                if hasta_adi and (not hasta_tahsilat.empty or not hasta_tedavi.empty):
                    issue = {
                        "hasta_id": hasta_id,
                        "hasta_adi": hasta_adi,
                        "hekim_sayisi": 0,
                        "toplam_tedavi": float(hasta_tedavi["TEDAVI_TOPLAM"].sum()),
                        "toplam_tahsilat": 0,
                        "toplam_fark": 0,
                        "sorun_tipi": "Normal",
                        "hekimler": []
                    }
                    
                    # Bu hastaya ait TÜM tahsilatları al
                    hasta_tum_tahsilat = tum_tahsilat_df[tum_tahsilat_df["HASTA_ID"].astype(str) == str(hasta_id)]
                    
                    # Tüm hekimleri topla
                    tedavi_hekimler = set(hasta_tedavi["DOKTOR_ID"].tolist())
                    tahsilat_hekimler = set(hasta_tum_tahsilat["DOKTOR_ID"].tolist())
                    tum_hekimler = tedavi_hekimler.union(tahsilat_hekimler)
                    
                    toplam_tahsilat_all = 0
                    
                    for hekim_id in tum_hekimler:
                        # Tedavi bilgileri
                        hekim_tedavi = hasta_tedavi[hasta_tedavi["DOKTOR_ID"] == hekim_id]
                        tedavi_tutari = float(hekim_tedavi["TEDAVI_TOPLAM"].iloc[0]) if not hekim_tedavi.empty else 0.0
                        tedavi_sayisi = int(hekim_tedavi["TEDAVI_SAYISI"].iloc[0]) if not hekim_tedavi.empty else 0
                        ilk_tedavi = str(hekim_tedavi["ILK_TEDAVI"].iloc[0]) if not hekim_tedavi.empty else "Tedavi Yok"
                        son_tedavi = str(hekim_tedavi["SON_TEDAVI"].iloc[0]) if not hekim_tedavi.empty else "Tedavi Yok"
                        
                        # Tahsilat bilgileri
                        hekim_tahsilat = hasta_tum_tahsilat[hasta_tum_tahsilat["DOKTOR_ID"] == hekim_id]
                        tahsilat_tutari = float(hekim_tahsilat["TOPLAM_TUTAR"].iloc[0]) if not hekim_tahsilat.empty else 0.0
                        tahsilat_sayisi = int(hekim_tahsilat["TAHSILAT_SAYISI"].iloc[0]) if not hekim_tahsilat.empty else 0
                        son_tahsilat_tarihi = str(hekim_tahsilat["SON_TAHSILAT_TARIHI"].iloc[0]) if not hekim_tahsilat.empty else "Tahsilat Yok"
                        
                        # Hekim adını önceden çekilmiş dict'ten al
                        if not hekim_tedavi.empty:
                            hekim_adi = hekim_tedavi["HEKIM_ADI"].iloc[0]
                        else:
                            hekim_adi = hekim_adlari_dict.get(hekim_id, f"Hekim ID: {hekim_id}")
                        
                        toplam_tahsilat_all += tahsilat_tutari
                        
                        # Sıfır olmayan tutarları göster
                        if tedavi_tutari > 0 or tahsilat_tutari > 0:
                            # Bu hekim seçilen hekimler arasında mı kontrol et
                            is_selected_doctor = str(hekim_id) in [str(d) for d in selected_doctors] if selected_doctors else False
                            
                            issue["hekimler"].append({
                                "hekim_adi": hekim_adi,
                                "hekim_id": hekim_id,
                                "is_selected": is_selected_doctor,
                                "tedavi_tutari": tedavi_tutari,
                                "tedavi_sayisi": tedavi_sayisi,
                                "tahsilat_tutari": tahsilat_tutari,
                                "tahsilat_sayisi": tahsilat_sayisi,
                                "son_tahsilat_tarihi": son_tahsilat_tarihi,
                                "fark": tahsilat_tutari - tedavi_tutari,
                                "ilk_tedavi": ilk_tedavi,
                                "son_tedavi": son_tedavi
                            })
                    
                    issue["hekim_sayisi"] = len(issue["hekimler"])
                    issue["toplam_tahsilat"] = toplam_tahsilat_all
                    issue["toplam_fark"] = toplam_tahsilat_all - issue["toplam_tedavi"]
                    
                    multi_doctor_issues.append(issue)

        logger.info(f"Tahsilat kayıt sayısı: {len(df)} | Çoklu hekim hasta sayısı: {len(multi_patients)}")

        # Özet hesaplamalar
        summary = {
            "toplam_tahsilat": f"{df['TUTAR'].sum():,.2f}",
            "toplam_kesinti": f"{df['KESINTI'].sum():,.2f}",
            "net_tahsilat": f"{df['NET_TAHSILAT'].sum():,.2f}",
            "toplam_prim": f"{df['HESAPLANAN_PRIM'].sum():,.2f}",
            "islem_sayisi": len(df)
        }

        # Şubeleri tekrar getir (template için)
        branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
        branches_df = pd.read_sql(branches_query, engine)
        branches = branches_df.to_dict('records')

        return render_template("tahsilatlar.html",
                             branches=branches,
                             table_data=df.to_dict("records"),
                             summary=summary,
                             multi_doctor_issues=multi_doctor_issues,
                             start_date=start_date,
                             end_date=end_date,
                             selected_branches=selected_branches,
                             selected_doctors=selected_doctors,
                             error_date=None)

    except Exception as e:
        logger.error(f"Tahsilatlar analizi hatası: {e}")
        engine = get_database_connection()
        branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
        branches_df = pd.read_sql(branches_query, engine)
        branches = branches_df.to_dict('records')
        flash("Analiz sırasında bir hata oluştu.", "danger")
        return render_template("tahsilatlar.html", 
                             branches=branches,
                             error_date=f"Hata: {str(e)}")
                             
@app.route("/ayarlar")
@login_required
def ayarlar():
    """Takvim ayarları sayfası"""
    return render_template("ayarlar.html")

@app.route("/api/calendar/settings", methods=["GET"])
@login_required
def get_calendar_settings():
    """Kullanıcının takvim ayarlarını getir"""
    try:
        username = session.get("username")
        
        # USERS dictionary'sinden kullanıcı ayarlarını oku
        user_data = USERS.get(username, {})
        user_settings = user_data.get('calendar_settings', {})
        
        # Varsayılan ayarlar
        default_settings = {
            "defaultView": "dayGridMonth",
            "startTime": "08:00",
            "endTime": "18:00",
            "slotDuration": "30",
            "showWeekends": True,
            "showAllDay": False,
            "appointmentColor": "#0d6efd",
            "urgentColor": "#dc3545",
            "completedColor": "#198754",
            "cancelledColor": "#6c757d",
            "language": "tr",
            "weekStart": "1",
            "timeFormat": "24",
            "dateFormat": "DD.MM.YYYY",
            "enableNotifications": True,
            "reminderMinutes": 30,
            "branchColors": []
        }
        
        # Varsayılan ile kullanıcı ayarlarını birleştir
        settings = {**default_settings, **user_settings}
        
        return jsonify({"success": True, "settings": settings})
        
    except Exception as e:
        logger.error(f"Takvim ayarları getirme hatası: {e}")
        return jsonify({"error": "Ayarlar yüklenemedi"}), 500

@app.route("/api/calendar/settings", methods=["POST"])
@login_required
def save_calendar_settings():
    """Kullanıcının takvim ayarlarını kaydet"""
    try:
        data = request.get_json() or {}
        username = session.get("username")
        
        if username not in USERS:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
        
        # USERS dictionary'sine ayarları ekle
        USERS[username]['calendar_settings'] = data
        
        # config.py'ye kaydet
        success = save_users_to_config()
        if success:
            return jsonify({"success": True, "message": "Takvim ayarları kaydedildi"})
        else:
            return jsonify({"error": "Ayarlar kaydedilemedi"}), 500
        
    except Exception as e:
        logger.error(f"Takvim ayarları kaydetme hatası: {e}")
        return jsonify({"error": "Ayarlar kaydedilemedi"}), 500

@app.route("/api/calendar/reset", methods=["POST"])
@login_required
def reset_calendar_settings():
    """Takvim ayarlarını varsayılana döndür"""
    try:
        username = session.get("username")
        
        if username in USERS:
            # Ayarları sil
            USERS[username].pop('calendar_settings', None)
            
            # config.py'ye kaydet
            success = save_users_to_config()
            if success:
                return jsonify({"success": True, "message": "Ayarlar varsayılana döndürüldü"})
            else:
                return jsonify({"error": "Ayarlar sıfırlanamadı"}), 500
        else:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
            
    except Exception as e:
        logger.error(f"Ayar sıfırlama hatası: {e}")
        return jsonify({"error": "Ayarlar sıfırlanamadı"}), 500
# app.py dosyasına eklenecek route'lar

@app.route("/api/prim/laboratuvar_gider_ekle", methods=["POST"])
@login_required
def laboratuvar_gider_ekle():
    """Laboratuvar gideri ekle"""
    try:
        data = request.get_json() or {}
        
        # Gerekli alanları kontrol et
        required_fields = ["tarih", "hasta_adi", "islem", "tutar"]
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"'{field}' alanı gereklidir"}), 400
        
        # Geçici ID ile kaydet (prim kaydetme sırasında prim_id güncellenecek)
        gider = {
            "id": f"temp_{datetime.now().timestamp()}",
            "tarih": data["tarih"],
            "hasta_adi": data["hasta_adi"],
            "hasta_id": data.get("hasta_id", ""),
            "islem": data["islem"],
            "tutar": float(data["tutar"])
        }
        
        return jsonify({"success": True, "gider": gider})
        
    except Exception as e:
        logger.error(f"Laboratuvar gider ekleme hatası: {e}")
        return jsonify({"error": "Gider eklenirken hata oluştu"}), 500

@app.route("/api/prim/implant_gider_ekle", methods=["POST"])
@login_required
def implant_gider_ekle():
    """İmplant gideri ekle"""
    try:
        data = request.get_json() or {}
        
        # Gerekli alanları kontrol et
        required_fields = ["tarih", "hasta_adi", "implant_markasi", "boy", "cap", "birim", "adet", "tutar"]
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"'{field}' alanı gereklidir"}), 400
        
        # Geçici ID ile kaydet
        gider = {
            "id": f"temp_{datetime.now().timestamp()}",
            "tarih": data["tarih"],
            "hasta_adi": data["hasta_adi"],
            "hasta_id": data.get("hasta_id", ""),
            "implant_markasi": data["implant_markasi"],
            "boy": data["boy"],
            "cap": data["cap"],
            "birim": data["birim"],
            "adet": int(data["adet"]),
            "tutar": float(data["tutar"])
        }
        
        return jsonify({"success": True, "gider": gider})
        
    except Exception as e:
        logger.error(f"İmplant gider ekleme hatası: {e}")
        return jsonify({"error": "Gider eklenirken hata oluştu"}), 500

@app.route("/api/prim/kaydet", methods=["POST"])
@login_required
def prim_kaydet_api():
    """Hesaplanan primi kaydet - CARİ ENTEGRASYONU + NET CİRO + HAK EDİŞ"""
    try:
        data = request.get_json() or {}
        
        prim_data = data.get('prim_data', {})
        tahsilat_detaylari = data.get('tahsilat_detaylari', [])
        diger_giderler = data.get('diger_giderler', [])
        laboratuvar_giderleri = data.get('laboratuvar_giderleri', [])
        implant_giderleri = data.get('implant_giderleri', [])
        
        # YENİ: Net Ciro ve Hak Ediş Eklemeleri
        net_ciro_eklemeleri = data.get('net_ciro_eklemeleri', [])
        hakedis_eklemeleri = data.get('hakedis_eklemeleri', [])
        
        # CARİ BİLGİSİ
        cari_id = data.get('cari_id')
        cari_eslestir = data.get('cari_eslestir', False)
        
        # Toplam gideri hesapla
        toplam_diger = sum(float(g.get('tutar', 0)) for g in diger_giderler)
        toplam_lab = sum(float(g.get('tutar', 0)) for g in laboratuvar_giderleri)
        toplam_implant = sum(float(g.get('tutar', 0)) for g in implant_giderleri)
        
        prim_data['toplam_gider'] = toplam_diger + toplam_lab + toplam_implant
        
        # Veri doğrulama
        errors = validate_prim_data(prim_data, tahsilat_detaylari, diger_giderler)
        if errors:
            return jsonify({"error": ", ".join(errors)}), 400
        
        # Kullanıcı bilgisini ekle
        prim_data["olusturan_kullanici"] = session.get("username")
        
        # PRİMİ KAYDET (GÜNCELLENMİŞ - yeni parametreler eklendi)
        prim_id = prim_db.prim_hesaplama_kaydet(
            prim_data, 
            tahsilat_detaylari, 
            diger_giderler,
            laboratuvar_giderleri,
            implant_giderleri,
            net_ciro_eklemeleri,  # YENİ
            hakedis_eklemeleri    # YENİ
        )
        
        # CARİ İŞLEMLERİ
        if cari_id:
            # Cari hesaba ALACAK kaydı ekle
            hareket_data = {
                'cari_id': cari_id,
                'hareket_tipi': 'prim_alacak',
                'prim_id': prim_id,
                'tarih': datetime.now().strftime('%Y-%m-%d'),
                'aciklama': f"Prim #{prim_id} - {prim_data['doktor_adi']} - {prim_data['sube_adi']} ({prim_data['donem_baslangic']} / {prim_data['donem_bitis']})",
                'alacak': prim_data['hesaplanan_prim'],
                'borc': 0,
                'olusturan_kullanici': session.get('username')
            }
            
            success, new_bakiye = cari_db.cari_hareket_ekle(hareket_data)
            
            if not success:
                logger.warning(f"Cari hareketi eklenemedi, ama prim kaydedildi: {prim_id}")
            
            # Eşleştirmeyi kaydet (checkbox işaretliyse)
            if cari_eslestir:
                cari_db.hekim_cari_eslestir(
                    cari_id,
                    prim_data['doktor_id'],
                    prim_data['doktor_adi'],
                    prim_data['sube_id'],
                    prim_data['sube_adi']
                )
        
        return jsonify({
            "success": True,
            "message": "Prim hesaplaması başarıyla kaydedildi",
            "prim_id": prim_id,
            "cari_islendi": bool(cari_id),
            "net_ciro_sayisi": len(net_ciro_eklemeleri),
            "hakedis_sayisi": len(hakedis_eklemeleri)
        })
        
    except Exception as e:
        logger.error(f"Prim kaydetme hatası: {e}")
        return jsonify({"error": "Prim kaydedilirken hata oluştu"}), 500
# Cari Yönetimi Route'ları

@app.route("/cari_yonetimi")
@login_required
def cari_yonetimi():
    """Cari hesaplar yönetim sayfası"""
    if session.get("role") != "admin":
        flash("Bu sayfaya erişim yetkiniz yok", "danger")
        return redirect(url_for("home"))
    
    try:
        # Şubeleri getir (Hekim eşleştirmede kullanılacak)
        query = "SELECT CARI_ID AS id, UNVANI AS name FROM subeler WHERE SILINDI = :silindi ORDER BY UNVANI"
        df = execute_query(query, {"silindi": "false"})
        branches = df.to_dict("records")
        
        return render_template("cari_yonetimi.html", branches=branches)
        
    except Exception as e:
        logger.error(f"Cari Yönetimi sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu: " + str(e), "danger")
        return redirect(url_for("home"))

@app.route("/api/cari/sil", methods=["POST"])
@admin_required
def cari_sil_api():
    """Cari hesabı silme - GELİŞTİRİLMİŞ VERSİYON"""
    try:
        data = request.get_json() or {}
        cari_id = data.get('cari_id')
        
        if not cari_id:
            return jsonify({"error": "Cari ID gerekli"}), 400
        
        # Gelişmiş silme fonksiyonunu kullan
        success, message = cari_db.cari_sil_gelismis(cari_id)
        
        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 400
            
    except Exception as e:
        logger.error(f"Cari silme API hatası: {e}")
        return jsonify({"error": "Cari silinirken hata oluştu"}), 500

       
@app.route("/api/cari/hareket_detay/<int:hareket_id>")
@login_required
def cari_hareket_detay_api(hareket_id):
    """Tek bir cari hareketin detayını getirir (Düzenleme Modalı için)"""
    try:
        hareket = cari_db.cari_hareket_detay_getir(hareket_id)
        if not hareket:
            return jsonify({"success": False, "error": "Hareket bulunamadı"}), 404
        return jsonify({"success": True, "data": hareket})
    except Exception as e:
        logger.error(f"Cari hareket detay API hatası: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/cari/hareket_duzelt", methods=["POST"])
@admin_required
def cari_hareket_duzelt_api():
    """Cari hareketi düzeltir ve bakiyeyi yeniden hesaplar"""
    try:
        data = request.get_json() or {}
        required_fields = ["id", "cari_id", "tarih", "aciklama", "alacak", "borc"]
        if not validate_required_fields(data, required_fields)[0]:
            return jsonify({"error": "Eksik parametreler"}), 400
            
        success, message = cari_db.cari_hareket_duzelt(
            data['id'], data['cari_id'], data['tarih'], data['aciklama'], float(data['alacak']), float(data['borc'])
        )

        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 500
            
    except Exception as e:
        logger.error(f"Cari hareket düzeltme hatası: {e}")
        return jsonify({"error": "Hareket düzeltilirken hata oluştu: " + str(e)}), 500

@app.route("/api/cari/hareket_sil", methods=["POST"])
@admin_required
def cari_hareket_sil_api():
    """Cari hareketi siler ve bakiyeyi yeniden hesaplar"""
    try:
        data = request.get_json() or {}
        hareket_id = data.get('id')
        cari_id = data.get('cari_id')
        
        if not hareket_id or not cari_id:
            return jsonify({"error": "Hareket ID ve Cari ID gerekli"}), 400
            
        success, message = cari_db.cari_hareket_sil(hareket_id, cari_id)

        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 500
            
    except Exception as e:
        logger.error(f"Cari hareket silme hatası: {e}")
        return jsonify({"error": "Hareket silinirken hata oluştu: " + str(e)}), 500

@app.route("/api/cari/ekstre_yazdir/<int:cari_id>")
@login_required
def cari_ekstre_yazdir(cari_id):
    """Cari hesap hareketlerini çekerek yazdırma HTML'ini döndürür"""
    try:
        # database.py'ye eklenen cari_detay_getir fonksiyonunu kullanırız
        cari_data = cari_db.cari_detay_getir(cari_id) 
        hareketler = cari_db.cari_hareket_listele(cari_id)
        
        if not cari_data:
            return "Cari hesap bulunamadı", 404
        
        ekstre_data = {
            'cari': cari_data,
            'hareketler': hareketler,
            'rapor_tarihi': datetime.now().strftime('%d.%m.%Y %H:%M')
        }
        
        # NOT: cari_ekstre_yazdir.html dosyası oluşturulmalıdır (Adım 4)
        return render_template("cari_ekstre_yazdir.html", detay=ekstre_data)
        
    except Exception as e:
        logger.error(f"Cari ekstre yazdırma hatası: {e}")
        return f"Ekstre oluşturulurken hata oluştu: {str(e)}", 500
@app.route("/api/cari/guncelle", methods=["POST"])
@admin_required
def cari_guncelle_api():
    """Mevcut cari hesabı güncelle"""
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz işlem"}), 403
    
    try:
        data = request.get_json() or {}
        cari_id = data.get('id')
        
        if not cari_id or not data.get('cari_kodu') or not data.get('cari_adi'):
            return jsonify({"error": "Cari ID, kodu ve adı gereklidir"}), 400
        
        success, message = cari_db.cari_guncelle(cari_id, data)
        
        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 500
            
    except Exception as e:
        logger.error(f"Cari güncelleme hatası: {e}")
        return jsonify({"error": "Cari güncellenemedi: " + str(e)}), 500

@app.route("/api/cari/eslestirme_liste/<int:cari_id>")
@login_required
def eslestirme_liste_api(cari_id):
    """Bir cariye ait tüm hekim eşleştirmelerini listele"""
    try:
        eslestirmeler = cari_db.eslestirme_listele(cari_id)
        return jsonify({"success": True, "data": eslestirmeler})
    except Exception as e:
        logger.error(f"Eşleştirme listeleme hatası: {e}")
        return jsonify({"error": "Eşleştirmeler yüklenemedi"}), 500

@app.route("/api/cari/eslestirme_sil", methods=["POST"])
@admin_required
def eslestirme_sil_api():
    """Hekim-cari eşleştirmesini sil"""
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz işlem"}), 403
        
    try:
        data = request.get_json() or {}
        eslestirme_id = data.get('id')
        
        if not eslestirme_id:
            return jsonify({"error": "Eşleştirme ID gerekli"}), 400
            
        success, message = cari_db.eslestirme_sil(eslestirme_id)
        
        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 500
            
    except Exception as e:
        logger.error(f"Eşleştirme silme hatası: {e}")
        return jsonify({"error": "Eşleştirme silinemedi"}), 500

@app.route("/api/cari/hareket_liste/<int:cari_id>")
@login_required
def cari_hareket_liste_api(cari_id):
    """Bir cariye ait tüm hareketleri listele"""
    try:
        hareketler = cari_db.cari_hareket_listele(cari_id)
        return jsonify({"success": True, "data": hareketler})
    except Exception as e:
        logger.error(f"Cari hareket listeleme hatası: {e}")
        return jsonify({"error": "Cari hareketler yüklenemedi"}), 500
        
@app.route("/api/cari/liste")
@login_required
def cari_liste_api():
    """Cari listesini getir (filtreli)"""
    try:
        cari_turu = request.args.get("cari_turu")
        alt_turu = request.args.get("alt_turu")
        cari_grup = request.args.get("cari_grup")
        
        cariler = cari_db.cari_listele(
            cari_turu=cari_turu if cari_turu else None,
            alt_turu=alt_turu if alt_turu else None
        )
        
        return jsonify({"success": True, "data": cariler})
        
    except Exception as e:
        logger.error(f"Cari listeleme hatası: {e}")
        return jsonify({"error": "Cari listesi yüklenemedi"}), 500
        

@app.route("/api/cari/turler")
@login_required
def cari_turler_api():
    """Veritabanındaki mevcut cari_turu ve alt_turu değerlerini döner"""
    try:
        turler = cari_db.cari_turleri_getir()
        return jsonify({"success": True, "data": turler})
    except Exception as e:
        logger.error(f"Cari türleri API hatası: {e}")
        return jsonify({"error": "Türler yüklenemedi"}), 500


@app.route("/api/cari/ekle", methods=["POST"])
@login_required
def cari_ekle_api():
    """Yeni cari ekle"""
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz işlem"}), 403
    
    try:
        data = request.get_json() or {}
        
        # Gerekli alanları kontrol et
        if not data.get('cari_kodu') or not data.get('cari_adi'):
            return jsonify({"error": "Cari kodu ve adı gereklidir"}), 400
        
        cari_id = cari_db.cari_ekle(data)
        
        return jsonify({
            "success": True,
            "message": "Cari hesap oluşturuldu",
            "cari_id": cari_id
        })
        
    except Exception as e:
        logger.error(f"Cari ekleme hatası: {e}")
        return jsonify({"error": "Cari eklenemedi: " + str(e)}), 500

@app.route("/api/cari/hekim_bul", methods=["POST"])
@login_required
def cari_hekim_bul_api():
    """Hekim-şube için cari bul"""
    try:
        data = request.get_json() or {}
        
        doktor_id = data.get('doktor_id')
        sube_id = data.get('sube_id')
        
        if not doktor_id or not sube_id:
            return jsonify({"success": False, "cari": None})
        
        cari = cari_db.cari_bul_hekim_sube(doktor_id, sube_id)
        
        return jsonify({
            "success": True,
            "cari": cari
        })
        
    except Exception as e:
        logger.error(f"Cari bulma hatası: {e}")
        return jsonify({"success": False, "cari": None})

# app.py dosyası içinde, "/api/cari/eslestir" rotası
@app.route("/api/cari/eslestir", methods=["POST"])
@login_required
def cari_eslestir_api():
    """Hekim-şube ile cari eşleştir (database.py ile uyumlu ad kullanıldı)"""
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz işlem"}), 403
    
    try:
        data = request.get_json() or {}
        
        # database.py'deki hekim_cari_eslestir fonksiyonunu çağırıyoruz.
        success, message = cari_db.hekim_cari_eslestir(
            data['cari_id'],
            data['doktor_id'],
            data['doktor_adi'],
            data['sube_id'],
            data['sube_adi']
        )
        
        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 400
        
    except Exception as e:
        logger.error(f"Eşleştirme hatası: {e}")
        # Hata mesajı, veritabanından gelen asıl hatayı da içerdiği için, daha genel bir mesajla döneriz.
        return jsonify({"error": "Eşleştirme sırasında sunucu hatası oluştu. Logları kontrol edin."}), 500
        
@app.route("/api/cari/hareket_ekle", methods=["POST"])
@login_required
def cari_hareket_ekle_api():
    """Cari hesaba hareket ekle (ödeme vb.)"""
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz işlem"}), 403
    
    try:
        data = request.get_json() or {}
        
        data['olusturan_kullanici'] = session.get('username')
        
        success, new_bakiye = cari_db.cari_hareket_ekle(data)
        
        if success:
            return jsonify({
                "success": True,
                "message": "Hareket kaydedildi",
                "yeni_bakiye": new_bakiye
            })
        else:
            return jsonify({"error": "Hareket eklenemedi"}), 500
        
    except Exception as e:
        logger.error(f"Hareket ekleme hatası: {e}")
        return jsonify({"error": str(e)}), 500
# app.py dosyasına eklenecek route'lar
# from database import prim_db, cari_db satırını şu şekilde güncelleyin:
# from database import prim_db, cari_db, personel_db

# ==================== PERSONEL YÖNETİMİ ROUTES ====================

@app.route("/personel")
@login_required
def personel():
    """Personel yönetimi ana sayfası"""
    if session.get("role") != "admin":
        flash("Bu sayfaya erişim yetkiniz yok", "danger")
        return redirect(url_for("home"))
    
    try:
        # Şubeleri getir
        query = "SELECT CARI_ID AS id, UNVANI AS name FROM subeler WHERE SILINDI = :silindi ORDER BY UNVANI"
        df = execute_query(query, {"silindi": "false"})
        branches = df.to_dict("records")
        
        return render_template("personel.html", branches=branches)
        
    except Exception as e:
        logger.error(f"Personel sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu", "danger")
        return redirect(url_for("home"))

# ==================== PERSONEL CRUD API ====================

@app.route("/api/personel/liste")
@login_required
def personel_liste_api():
    """Personel listesini getir"""
    try:
        sadece_aktif = request.args.get("sadece_aktif", "true").lower() == "true"
        sube_id = request.args.get("sube_id")
        
        personeller = personel_db.personel_listele(sadece_aktif=sadece_aktif, sube_id=sube_id)
        
        return jsonify({
            "success": True,
            "data": personeller
        })
        
    except Exception as e:
        logger.error(f"Personel listeleme hatası: {e}")
        return jsonify({"error": "Personel listesi yüklenemedi"}), 500

@app.route("/api/personel/detay/<int:personel_id>")
@login_required
def personel_detay_api(personel_id):
    """Personel detayını getir"""
    try:
        detay = personel_db.personel_detay_getir(personel_id)
        
        if not detay:
            return jsonify({"error": "Personel bulunamadı"}), 404
        
        # Fotoğrafı base64'e çevir (eğer varsa)
        if detay['personel'].get('fotograf'):
            import base64
            detay['personel']['fotograf_base64'] = base64.b64encode(detay['personel']['fotograf']).decode('utf-8')
            detay['personel']['fotograf'] = None  # BLOB'u frontend'e göndermiyoruz
        
        return jsonify({
            "success": True,
            "data": detay
        })
        
    except Exception as e:
        logger.error(f"Personel detay getirme hatası: {e}")
        return jsonify({"error": "Personel detayı yüklenemedi"}), 500

@app.route("/api/personel/ekle", methods=["POST"])
@admin_required
def personel_ekle_api():
    """Yeni personel ekle + Maaş bilgileri"""
    try:
        data = request.get_json() or {}
        
        # Gerekli alanları kontrol et
        required_fields = ["tc_kimlik", "ad", "soyad", "ise_baslama_tarihi"]
        valid, error_msg = validate_required_fields(data, required_fields)
        if not valid:
            return jsonify({"error": error_msg}), 400
        
        # TC Kimlik format kontrolü (11 haneli)
        if len(data['tc_kimlik']) != 11 or not data['tc_kimlik'].isdigit():
            return jsonify({"error": "TC Kimlik 11 haneli sayı olmalıdır"}), 400
        
        # TC kimlik tekrar kontrolü
        import sqlite3
        conn = sqlite3.connect(personel_db.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, ad, soyad, calisma_durumu FROM personel WHERE tc_kimlik = ?", (data['tc_kimlik'],))
        existing = cursor.fetchone()
        conn.close()
        
        if existing:
            durum = "Aktif" if existing[3] == 'aktif' else "Pasif"
            return jsonify({
                "error": f"Bu TC kimlik numarası zaten kayıtlı!\n\nMevcut Personel: {existing[1]} {existing[2]}\nDurum: {durum}\nID: {existing[0]}\n\nLütfen farklı bir TC kimlik numarası girin."
            }), 400
        
        # Pozisyon kontrolü - Hekim mi?
        is_hekim = (data.get('pozisyon') == 'hekim')
        
        # Hekim değilse maaş bilgileri zorunlu
        if not is_hekim:
            maas_bilgileri = data.get('maas_bilgileri', {})
            if not maas_bilgileri.get('brut_maas') or not maas_bilgileri.get('net_maas'):
                return jsonify({"error": "Hekim olmayan personel için maaş bilgileri zorunludur"}), 400
        
        # Fotoğraf varsa base64'ten çevir
        if data.get('fotograf_base64'):
            import base64
            data['fotograf'] = base64.b64decode(data['fotograf_base64'])
        
        # Cari oluşturulsun mu?
        cari_olustur = data.get('cari_olustur', True)
        
        # Personeli ekle
        personel_id, cari_id = personel_db.personel_ekle(data, cari_olustur=cari_olustur)
        
        # Maaş bilgilerini kaydet (hekim değilse)
        if not is_hekim and data.get('maas_bilgileri'):
            maas_data = data['maas_bilgileri']
            maas_data['personel_id'] = personel_id
            
            success = personel_db.maas_bilgisi_tanimla(personel_id, maas_data)
            if not success:
                logger.warning(f"Personel eklendi ama maaş tanımlanamadı: ID {personel_id}")
        
        return jsonify({
            "success": True,
            "message": "Personel başarıyla eklendi" + (" ve maaş bilgileri kaydedildi" if not is_hekim else ""),
            "personel_id": personel_id,
            "cari_id": cari_id
        })
        
    except sqlite3.IntegrityError as e:
        error_msg = str(e)
        
        if "UNIQUE constraint failed: personel.tc_kimlik" in error_msg:
            return jsonify({
                "error": "Bu TC Kimlik numarası zaten kayıtlı! Lütfen kontrol edin."
            }), 400
        elif "UNIQUE constraint failed: cari_hesaplar.cari_kodu" in error_msg:
            return jsonify({
                "error": "Bu personel için cari hesap zaten mevcut."
            }), 400
        else:
            logger.error(f"Veritabanı hatası: {e}")
            return jsonify({
                "error": "Veritabanı hatası: Kayıt sırasında bir çakışma oluştu."
            }), 400
            
    except Exception as e:
        logger.error(f"Personel ekleme hatası: {e}")
        return jsonify({"error": f"Personel eklenirken hata oluştu: {str(e)}"}), 500

@app.route("/api/personel/maas/tanimla", methods=["POST"])
@admin_required
def personel_maas_tanimla():
    """Personel maaş bilgisi tanımla"""
    try:
        data = request.get_json() or {}
        
        required_fields = ["personel_id", "brut_maas", "net_maas", "baslangic_tarihi"]
        valid, error_msg = validate_required_fields(data, required_fields)
        if not valid:
            return jsonify({"error": error_msg}), 400
        
        success = personel_db.maas_bilgisi_tanimla(data['personel_id'], data)
        
        if success:
            return jsonify({
                "success": True,
                "message": "Maaş bilgisi tanımlandı"
            })
        else:
            return jsonify({"error": "Maaş tanımlama başarısız"}), 500
            
    except Exception as e:
        logger.error(f"Maaş tanımlama hatası: {e}")
        return jsonify({"error": "Maaş tanımlanamadı"}), 500

# app.py içinde mevcut /api/personel/maas/hesapla route'unu BUL ve DEĞİŞTİR
# Yaklaşık satır 1100 civarında olmalı

@app.route("/api/personel/maas/hesapla", methods=["POST"])
@login_required
def maas_hesapla_api():
    """Maaş hesaplama - Yardımlar Otomatik Eklenmiş"""
    try:
        data = request.get_json() or {}
        
        personel_id = data.get('personel_id')
        donem_ay = data.get('donem_ay')
        donem_yil = data.get('donem_yil')
        
        if not all([personel_id, donem_ay, donem_yil]):
            return jsonify({"error": "Eksik parametreler"}), 400
        
        # Personel maaş bilgilerini getir
        detay = personel_db.personel_detay_getir(personel_id)
        
        if not detay or not detay.get('maas'):
            return jsonify({"error": "Personel maaş bilgisi bulunamadı"}), 404
        
        maas_bilgisi = detay['maas']
        
        # Temel maaş bilgileri
        brut_maas = float(maas_bilgisi.get('brut_maas', 0))
        net_maas = float(maas_bilgisi.get('net_maas', 0))
        
        # OTOMATIK YARDIMLAR
        yol_yardimi = float(maas_bilgisi.get('yol_yardimi', 0))
        yemek_yardimi = float(maas_bilgisi.get('yemek_yardimi', 0))
        cocuk_yardimi = float(maas_bilgisi.get('cocuk_yardimi', 0))
        diger_odenekler = float(maas_bilgisi.get('diger_odenekler', 0))
        
        toplam_sabit_yardimlar = yol_yardimi + yemek_yardimi + cocuk_yardimi + diger_odenekler
        
        # Ücretsiz izin kesintisi
        ucretsiz_izin_gun = float(data.get('ucretsiz_izin_gun', 0))
        ucretsiz_izin_kesinti = (net_maas / 30) * ucretsiz_izin_gun if ucretsiz_izin_gun > 0 else 0
        
        # Fazla mesai (Saatlik ücret = Net Maaş / 220 saat, %50 fazlası)
        fazla_mesai_saat = float(data.get('fazla_mesai_saat', 0))
        saatlik_ucret = net_maas / 220 if net_maas > 0 else 0
        fazla_mesai_ucret = (saatlik_ucret * 1.5) * fazla_mesai_saat if fazla_mesai_saat > 0 else 0
        
        # Manuel eklemeler
        prim = float(data.get('prim', 0))
        bonus = float(data.get('bonus', 0))
        
        # TOPLAM HESAPLAMA
        odenecek_tutar = (
            net_maas 
            + toplam_sabit_yardimlar  # OTOMATIK EKLENEN
            - ucretsiz_izin_kesinti 
            + fazla_mesai_ucret 
            + prim 
            + bonus
        )
        
        hesaplama = {
            'brut_maas': brut_maas,
            'net_maas': net_maas,
            
            # YARDIMLAR (OTOMATIK)
            'yol_yardimi': yol_yardimi,
            'yemek_yardimi': yemek_yardimi,
            'cocuk_yardimi': cocuk_yardimi,
            'diger_odenekler': diger_odenekler,
            'toplam_sabit_yardimlar': toplam_sabit_yardimlar,
            
            # KESİNTİLER
            'ucretsiz_izin_gun': ucretsiz_izin_gun,
            'ucretsiz_izin_kesinti': ucretsiz_izin_kesinti,
            
            # EK ÖDEMELER
            'fazla_mesai_saat': fazla_mesai_saat,
            'fazla_mesai_ucret': fazla_mesai_ucret,
            'prim': prim,
            'bonus': bonus,
            
            # TOPLAM
            'odenecek_tutar': odenecek_tutar
        }
        
        return jsonify({
            "success": True,
            "hesaplama": hesaplama
        })
        
    except Exception as e:
        logger.error(f"Maaş hesaplama hatası: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Hesaplama yapılamadı: {str(e)}"}), 500

@app.route("/api/personel/maas/kaydet", methods=["POST"])
@admin_required
def personel_maas_kaydet():
    """Maaş ödemesi kaydet ve cari hesaba işle"""
    try:
        data = request.get_json() or {}
        data['olusturan_kullanici'] = session.get('username')
        
        # 1. Maaş hesaplamasını kaydet
        maas_id = personel_db.maas_odeme_kaydet(data)
        
        # 2. Cari hesaba ALACAK olarak işle
        if data.get('cari_id'):
            cari_hareket = {
                'cari_id': data['cari_id'],
                'hareket_tipi': 'maas_alacak',
                'tarih': data.get('odeme_tarihi', datetime.now().strftime('%Y-%m-%d')),
                'aciklama': f"Maaş Hesaplaması - {data['donem_ay']}/{data['donem_yil']}",
                'alacak': data['odenecek_tutar'],
                'borc': 0,
                'olusturan_kullanici': session.get('username')
            }
            
            success, new_bakiye = cari_db.cari_hareket_ekle(cari_hareket)
            
            if not success:
                logger.warning(f"Maaş kaydedildi ama cari işlenemedi: Maaş ID {maas_id}")
        
        return jsonify({
            "success": True,
            "message": "Maaş hesaplaması kaydedildi ve cari hesaba alacak olarak işlendi",
            "maas_id": maas_id
        })
        
    except Exception as e:
        logger.error(f"Maaş kaydetme hatası: {e}")
        return jsonify({"error": f"Maaş kaydedilemedi: {str(e)}"}), 500
        
@app.route("/api/personel/maas/liste")
@login_required
def personel_maas_liste():
    """Maaş ödemelerini listele (filtreli)"""
    try:
        donem_ay = request.args.get("donem_ay")
        donem_yil = request.args.get("donem_yil")
        personel_id = request.args.get("personel_id")
        cari_turu = request.args.get("cari_turu")
        alt_turu = request.args.get("alt_turu")

        maaslar = personel_db.maas_listele(
            donem_ay=int(donem_ay) if donem_ay else None,
            donem_yil=int(donem_yil) if donem_yil else None,
            personel_id=int(personel_id) if personel_id else None,
            cari_turu=cari_turu if cari_turu else None,
            alt_turu=alt_turu if alt_turu else None
        )
        
        return jsonify({
            "success": True,
            "data": maaslar
        })
        
    except Exception as e:
        logger.error(f"Maaş listeleme hatası: {e}")
        return jsonify({"error": "Maaş listesi yüklenemedi"}), 500


# ==================== İZİN YÖNETİMİ ====================

@app.route("/api/personel/izin/ekle", methods=["POST"])
@admin_required
def personel_izin_ekle():
    """İzin kaydı ekle"""
    try:
        data = request.get_json() or {}
        
        conn = sqlite3.connect(personel_db.db_path)
        cursor = conn.cursor()
        
        # Gün sayısını hesapla
        from datetime import datetime
        baslangic = datetime.strptime(data['baslangic_tarihi'], '%Y-%m-%d')
        bitis = datetime.strptime(data['bitis_tarihi'], '%Y-%m-%d')
        gun_sayisi = (bitis - baslangic).days + 1
        
        cursor.execute('''
            INSERT INTO personel_izin
            (personel_id, izin_tipi, baslangic_tarihi, bitis_tarihi, gun_sayisi, 
             aciklama, onay_durumu, onaylayan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['personel_id'],
            data['izin_tipi'],
            data['baslangic_tarihi'],
            data['bitis_tarihi'],
            gun_sayisi,
            data.get('aciklama'),
            data.get('onay_durumu', 'onaylandi'),
            session.get('username')
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "İzin kaydı eklendi",
            "gun_sayisi": gun_sayisi
        })
        
    except Exception as e:
        logger.error(f"İzin ekleme hatası: {e}")
        return jsonify({"error": "İzin eklenemedi"}), 500

@app.route("/api/personel/izin/liste/<int:personel_id>")
@login_required
def personel_izin_liste(personel_id):
    """Personel izinlerini listele"""
    try:
        conn = sqlite3.connect(personel_db.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM personel_izin 
            WHERE personel_id = ? 
            ORDER BY baslangic_tarihi DESC
        ''', (personel_id,))
        
        izinler = cursor.fetchall()
        conn.close()
        
        return jsonify({
            "success": True,
            "data": [dict(izin) for izin in izinler]
        })
        
    except Exception as e:
        logger.error(f"İzin listeleme hatası: {e}")
        return jsonify({"error": "İzinler listelenemedi"}), 500

# ==================== DOKÜMAN YÖNETİMİ ====================

@app.route("/api/personel/dokuman/ekle", methods=["POST"])
@admin_required
def personel_dokuman_ekle():
    """Personel dokümanı ekle"""
    try:
        data = request.get_json() or {}
        
        conn = sqlite3.connect(personel_db.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO personel_dokuman
            (personel_id, dokuman_tipi, dosya_adi, dosya_yolu, onay_tarihi, 
             gecerlilik_tarihi, notlar)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['personel_id'],
            data['dokuman_tipi'],
            data.get('dosya_adi'),
            data.get('dosya_yolu'),
            data.get('onay_tarihi'),
            data.get('gecerlilik_tarihi'),
            data.get('notlar')
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Doküman kaydedildi"})
        
    except Exception as e:
        logger.error(f"Doküman ekleme hatası: {e}")
        return jsonify({"error": "Doküman eklenemedi"}), 500
        
@app.route("/maas_yonetimi")
@login_required
def maas_yonetimi():
    """Maaş yönetimi sayfası"""
    if session.get("role") != "admin":
        flash("Bu sayfaya erişim yetkiniz yok", "danger")
        return redirect(url_for("home"))
    
    return render_template("maas_yonetimi.html")
@app.route("/api/personel/fotograf_yukle", methods=["POST"])
@admin_required
def personel_fotograf_yukle():
    """Personel fotoğrafı yükle"""
    try:
        if 'fotograf' not in request.files:
            return jsonify({"error": "Fotoğraf dosyası bulunamadı"}), 400
        
        file = request.files['fotograf']
        personel_id = request.form.get('personel_id')
        
        if not personel_id:
            return jsonify({"error": "Personel ID gerekli"}), 400
        
        if file.filename == '':
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        # Dosya tipini kontrol et
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({"error": "Sadece resim dosyaları yüklenebilir (PNG, JPG, JPEG, GIF)"}), 400
        
        # Dosya boyutu kontrolü (max 5MB)
        file.seek(0, 2)  # Dosya sonuna git
        file_size = file.tell()  # Boyutu al
        file.seek(0)  # Başa dön
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return jsonify({"error": "Dosya boyutu 5MB'dan küçük olmalıdır"}), 400
        
        # Dosyayı oku
        fotograf_data = file.read()
        
        # Veritabanına kaydet
        import sqlite3
        conn = sqlite3.connect(personel_db.db_path)
        cursor = conn.cursor()
        cursor.execute("UPDATE personel SET fotograf = ? WHERE id = ?", (fotograf_data, personel_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Fotoğraf başarıyla yüklendi"
        })
        
    except Exception as e:
        logger.error(f"Fotoğraf yükleme hatası: {e}")
        return jsonify({"error": "Fotoğraf yüklenemedi"}), 500
        
@app.route("/api/personel/guncelle", methods=["POST"])
@admin_required
def personel_guncelle_api():
    """Personel bilgilerini güncelle"""
    try:
        data = request.get_json() or {}
        personel_id = data.get('id')
        
        if not personel_id:
            return jsonify({"error": "Personel ID gerekli"}), 400
        
        # TC Kimlik kontrolü (güncelleme sırasında)
        tc_kimlik = data.get('tc_kimlik')
        if tc_kimlik:
            import sqlite3
            conn = sqlite3.connect(personel_db.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM personel WHERE tc_kimlik = ? AND id != ?", (tc_kimlik, personel_id))
            existing = cursor.fetchone()
            conn.close()
            
            if existing:
                return jsonify({"error": "Bu TC kimlik numarası başka bir personele ait!"}), 400
        
        success = personel_db.personel_guncelle(personel_id, data)
        
        if success:
            return jsonify({
                "success": True,
                "message": "Personel bilgileri güncellendi"
            })
        else:
            return jsonify({"error": "Güncelleme başarısız"}), 500
            
    except Exception as e:
        logger.error(f"Personel güncelleme hatası: {e}")
        return jsonify({"error": "Güncelleme sırasında hata oluştu"}), 500 
@app.route("/api/cari/detay/<int:cari_id>")
@login_required
def cari_detay_api(cari_id):
    """Cari hesap detayını getir"""
    try:
        cari = cari_db.cari_detay_getir(cari_id)
        
        if not cari:
            return jsonify({"error": "Cari hesap bulunamadı"}), 404
        
        return jsonify({
            "success": True,
            "data": cari
        })
        
    except Exception as e:
        logger.error(f"Cari detay getirme hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": "Cari detay yüklenemedi: " + str(e)}), 500        
@app.route("/api/personel/maas/odeme_yap", methods=["POST"])
@admin_required
def personel_maas_odeme_yap():
    """Bekleyen maaşı öde (cari hesaptan borç düş)"""
    try:
        data = request.get_json() or {}
        
        maas_id = data.get('maas_id')
        odeme_yontemi = data.get('odeme_yontemi')
        odeme_tarihi = data.get('odeme_tarihi', datetime.now().strftime('%Y-%m-%d'))
        
        if not maas_id or not odeme_yontemi:
            return jsonify({"error": "Maaş ID ve ödeme yöntemi gerekli"}), 400
        
        # 1. Maaş kaydını getir
        import sqlite3
        conn = sqlite3.connect(personel_db.db_path, timeout=30.0)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM maas_odeme WHERE id = ?", (maas_id,))
        maas = cursor.fetchone()
        
        if not maas:
            conn.close()
            return jsonify({"error": "Maaş kaydı bulunamadı"}), 404
        
        # Sütun indeksleri
        cari_id_idx = 2
        donem_ay_idx = 3
        donem_yil_idx = 4
        odenecek_tutar_idx = 12
        odeme_durumu_idx = 15
        
        if maas[odeme_durumu_idx] == 'odendi':
            conn.close()
            return jsonify({"error": "Bu maaş zaten ödenmiş"}), 400
        
        cari_id = maas[cari_id_idx]
        odenecek_tutar = maas[odenecek_tutar_idx]
        
        # 2. Maaş durumunu güncelle
        cursor.execute("""
            UPDATE maas_odeme 
            SET odeme_durumu = 'odendi', 
                odeme_tarihi = ?,
                odeme_yontemi = ?
            WHERE id = ?
        """, (odeme_tarihi, odeme_yontemi, maas_id))
        
        conn.commit()
        conn.close()
        
        # 3. Cari hesaba BORÇ ekle (ayrı bir connection ile)
        if cari_id:
            import time
            time.sleep(0.1)  # Kısa bir bekleme ekle
            
            cari_hareket = {
                'cari_id': cari_id,
                'hareket_tipi': 'maas_odeme',
                'tarih': odeme_tarihi,
                'aciklama': f"Maaş Ödemesi ({odeme_yontemi}) - {maas[donem_ay_idx]}/{maas[donem_yil_idx]}",
                'alacak': 0,
                'borc': odenecek_tutar,
                'olusturan_kullanici': session.get('username')
            }
            
            success, new_bakiye = cari_db.cari_hareket_ekle(cari_hareket)
            
            if not success:
                logger.warning(f"Maaş ödendi olarak işaretlendi ama cari güncellenemedi")
                return jsonify({
                    "success": True,
                    "message": "Maaş ödendi olarak işaretlendi, ancak cari güncelleme başarısız. Lütfen manuel kontrol edin.",
                    "warning": True
                })
        
        return jsonify({
            "success": True,
            "message": "Maaş ödemesi tamamlandı"
        })
        
    except Exception as e:
        logger.error(f"Maaş ödeme hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": f"Ödeme hatası: {str(e)}"}), 500
@app.route("/download_tahsilatlar", methods=["POST"])
@login_required
def download_tahsilatlar():
    """Tahsilat analizini Excel olarak indir"""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        data = request.get_json() or {}
        
        start_date_raw = data.get("start_date")
        end_date_raw = data.get("end_date")
        selected_branches = data.get("selected_branches", [])
        selected_doctors = data.get("selected_doctors", [])
        
        # Tarih kontrolü
        start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_raw, "%Y-%m-%d").date()
        
        engine = get_database_connection()
        
        # Tahsilat verilerini çek (analyze_tahsilatlar ile aynı sorgu)
        branch_placeholders = ','.join([':branch_' + str(i) for i in range(len(selected_branches))])
        
        params = {
            'start_date': start_date,
            'end_date': end_date
        }
        
        for i, branch in enumerate(selected_branches):
            params[f'branch_{i}'] = branch
        
        tahsilat_query = f"""
        SELECT
            T1.TARIH,
            T1.ALACAK AS TUTAR,
            T1.HEDEF_ILGILI_DOKTOR_ID AS DOKTOR_ID,
            CONCAT(IFNULL(DR.ADI,''),' ',IFNULL(DR.SOYADI,'')) AS HEKIM_ADI,
            IFNULL(DR.PRIMYUZDE,0) AS PRIMYUZDE,
            CONCAT(IFNULL(H.ADI,''),' ',IFNULL(H.SOYADI,'')) AS HASTA_ADI,
            H.HASTA_ID,
            LOWER(IFNULL(OS.ADI,'bilinmeyen')) AS ODEME_SEKLI,
            IFNULL(SB.UNVANI,'Bilinmeyen') AS SUBE_ADI,
            T1.HAREKETTYPE
        FROM carihareket AS T1
        LEFT JOIN kartdoktor AS DR ON T1.HEDEF_ILGILI_DOKTOR_ID = DR.CARI_ID
        LEFT JOIN odeme_sekilleri AS OS ON T1.ISLEM_TIPI_ID = OS.ROWNO
        LEFT JOIN subeler AS SB ON T1.SUBE_ID = SB.CARI_ID
        LEFT JOIN karthasta AS H ON T1.KAYNAK_CARI_ID = H.HASTA_ID
        WHERE T1.SILINDI='False'
          AND T1.ALACAK > 0
          AND T1.HAREKETTYPE IN ('T','ST','CT')
          AND T1.TARIH BETWEEN :start_date AND :end_date
          AND T1.SUBE_ID IN ({branch_placeholders})
        """
        
        if selected_doctors:
            doctor_placeholders = ','.join([':doctor_' + str(i) for i in range(len(selected_doctors))])
            tahsilat_query += f" AND T1.HEDEF_ILGILI_DOKTOR_ID IN ({doctor_placeholders})"
            for i, doctor in enumerate(selected_doctors):
                params[f'doctor_{i}'] = doctor
        
        df = pd.read_sql(text(tahsilat_query), engine, params=params)
        
        # Kesinti hesaplamaları
        KESINTI_ORANLARI = {
            'nakit': 0,
            'pos': 0.10,
            'banka': 0.10,
            'çek': 0.10,
            'senet': 0,
            'senet tahsilatı': 0
        }
        
        df["ODEME_SEKLI"] = df["ODEME_SEKLI"].str.strip().str.lower()
        df["KESINTI_ORANI"] = df["ODEME_SEKLI"].map(KESINTI_ORANLARI).fillna(0)
        df["KESINTI"] = df["TUTAR"] * df["KESINTI_ORANI"]
        df["NET_TAHSILAT"] = df["TUTAR"] - df["KESINTI"]
        df["PRIM_YUZDE"] = df["PRIMYUZDE"] / 100
        df["HESAPLANAN_PRIM"] = df["NET_TAHSILAT"] * df["PRIM_YUZDE"]
        
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        
        # 1. Sayfa: Tahsilat Detayları
        ws1 = wb.active
        ws1.title = "Tahsilat Detayları"
        
        # Başlıklar
        headers = ["HASTA ADI", "HEKİM ADI", "ÖDEME YÖNTEMİ", "ŞUBE ADI", "TARİH", 
                  "TUTAR", "KESİNTİ", "NET TAHSİLAT", "PRİM %", "HESAPLANAN PRİM"]
        
        # Başlık stili
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Veri satırları
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            ws1.cell(row=row_num, column=1, value=row_data.HASTA_ADI).border = border
            ws1.cell(row=row_num, column=2, value=row_data.HEKIM_ADI).border = border
            ws1.cell(row=row_num, column=3, value=row_data.ODEME_SEKLI).border = border
            ws1.cell(row=row_num, column=4, value=row_data.SUBE_ADI).border = border
            ws1.cell(row=row_num, column=5, value=str(row_data.TARIH)).border = border
            ws1.cell(row=row_num, column=6, value=float(row_data.TUTAR)).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=6).border = border
            ws1.cell(row=row_num, column=7, value=float(row_data.KESINTI)).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=7).border = border
            ws1.cell(row=row_num, column=8, value=float(row_data.NET_TAHSILAT)).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=8).border = border
            ws1.cell(row=row_num, column=9, value=float(row_data.PRIMYUZDE)).number_format = '0.00%'
            ws1.cell(row=row_num, column=9).border = border
            ws1.cell(row=row_num, column=10, value=float(row_data.HESAPLANAN_PRIM)).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=10).border = border
        
        # Özet satırı ekle
        last_row = len(df) + 2
        ws1.cell(row=last_row, column=1, value="TOPLAM:").font = Font(bold=True)
        ws1.cell(row=last_row, column=6, value=float(df['TUTAR'].sum())).number_format = '#,##0.00'
        ws1.cell(row=last_row, column=6).font = Font(bold=True)
        ws1.cell(row=last_row, column=7, value=float(df['KESINTI'].sum())).number_format = '#,##0.00'
        ws1.cell(row=last_row, column=7).font = Font(bold=True)
        ws1.cell(row=last_row, column=8, value=float(df['NET_TAHSILAT'].sum())).number_format = '#,##0.00'
        ws1.cell(row=last_row, column=8).font = Font(bold=True)
        ws1.cell(row=last_row, column=10, value=float(df['HESAPLANAN_PRIM'].sum())).number_format = '#,##0.00'
        ws1.cell(row=last_row, column=10).font = Font(bold=True)
        
        # Sütun genişliklerini ayarla
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 10
        ws1.column_dimensions['J'].width = 15
        
        # 2. Sayfa: Çoklu Hekim Uyarıları
        # Çoklu hekim hastalarını tespit et (analyze_tahsilatlar ile aynı mantık)
        tum_tedavi_query = """
        SELECT
            T1.HASTA_ID,
            T1.DOKTOR_ID,
            COALESCE(CONCAT(DR.ADI,' ',DR.SOYADI), '') AS HEKIM_ADI,
            SUM(COALESCE(T1.TUTAR, T1.LISTETUTAR, 0)) AS TEDAVI_TOPLAM,
            COUNT(*) AS TEDAVI_SAYISI,
            MIN(T1.TARIH) AS ILK_TEDAVI,
            MAX(T1.TARIH) AS SON_TEDAVI,
            DR.SUBE_ID
        FROM tedavi AS T1
        LEFT JOIN kartdoktor AS DR ON T1.DOKTOR_ID = DR.CARI_ID
        WHERE (T1.SILINDI IS NULL OR T1.SILINDI != 'True')
          AND (T1.ISDELETED IS NULL OR T1.ISDELETED != 'True')
        GROUP BY T1.HASTA_ID, T1.DOKTOR_ID, HEKIM_ADI, DR.SUBE_ID
        """
        tum_tedavi_df = pd.read_sql(tum_tedavi_query, engine)
        
        tum_multi_patients = set(
            tum_tedavi_df.groupby("HASTA_ID")["DOKTOR_ID"].nunique()
            .loc[lambda x: x > 1].index.astype(str).tolist()
        )
        
        ana_analiz_hastalari = set(df["HASTA_ID"].astype(str).tolist())
        multi_patients = tum_multi_patients.intersection(ana_analiz_hastalari)
        
        if selected_doctors:
            selected_doctor_patients = set(
                tum_tedavi_df[tum_tedavi_df["DOKTOR_ID"].isin(selected_doctors)]["HASTA_ID"].astype(str).tolist()
            )
            multi_patients = multi_patients.intersection(selected_doctor_patients)
        
        selected_branch_patients = set(
            tum_tedavi_df[tum_tedavi_df["SUBE_ID"].isin(selected_branches)]["HASTA_ID"].astype(str).tolist()
        )
        multi_patients = multi_patients.intersection(selected_branch_patients)
        
        if multi_patients:
            ws2 = wb.create_sheet(title="Çoklu Hekim Uyarıları")
            
            # Başlıklar
            multi_headers = ["HASTA ADI", "HASTA ID", "HEKİM ADI", "DOKTOR ID", 
                           "TEDAVİ TUTARI", "TAHSİLAT TUTARI", "FARK", "SON TAHSİLAT TARİHİ"]
            
            for col_num, header in enumerate(multi_headers, 1):
                cell = ws2.cell(row=1, column=col_num, value=header)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Hasta detayları için veri topla
            tedavi_df = tum_tedavi_df[tum_tedavi_df["HASTA_ID"].astype(str).isin(multi_patients)]
            
            # Tüm tahsilatları al
            hasta_ids_list = list(multi_patients)
            hasta_placeholders = ','.join([':hasta_' + str(i) for i in range(len(hasta_ids_list))])
            
            tum_tahsilat_params = {}
            for i, pid in enumerate(hasta_ids_list):
                tum_tahsilat_params[f'hasta_{i}'] = pid
                
            tum_tahsilat_query = f"""
            SELECT
                T1.HEDEF_ILGILI_DOKTOR_ID AS DOKTOR_ID,
                H.HASTA_ID,
                SUM(T1.ALACAK) AS TOPLAM_TUTAR,
                COUNT(*) AS TAHSILAT_SAYISI,
                MAX(T1.TARIH) AS SON_TAHSILAT_TARIHI
            FROM carihareket AS T1
            LEFT JOIN karthasta AS H ON T1.KAYNAK_CARI_ID = H.HASTA_ID
            WHERE T1.SILINDI='False'
              AND T1.ALACAK > 0
              AND T1.HAREKETTYPE IN ('T','ST','CT')
              AND H.HASTA_ID IN ({hasta_placeholders})
            GROUP BY T1.HEDEF_ILGILI_DOKTOR_ID, H.HASTA_ID
            """
            tum_tahsilat_df = pd.read_sql(text(tum_tahsilat_query), engine, params=tum_tahsilat_params)
            
            # Hasta adlarını toplu çek
            tum_hasta_adlari_query = f"""
            SELECT HASTA_ID, CONCAT(IFNULL(ADI,''), ' ', IFNULL(SOYADI,'')) AS HASTA_ADI 
            FROM karthasta 
            WHERE HASTA_ID IN ({hasta_placeholders})
            """
            hasta_adlari_df = pd.read_sql(text(tum_hasta_adlari_query), engine, params=tum_tahsilat_params)
            hasta_adlari_dict = dict(zip(hasta_adlari_df["HASTA_ID"].astype(str), hasta_adlari_df["HASTA_ADI"]))
            
            # Her hasta için satır ekle
            current_row = 2
            for hasta_id in multi_patients:
                hasta_tedavi = tedavi_df[tedavi_df["HASTA_ID"].astype(str) == str(hasta_id)]
                hasta_tahsilat = tum_tahsilat_df[tum_tahsilat_df["HASTA_ID"].astype(str) == str(hasta_id)]
                
                hasta_adi = hasta_adlari_dict.get(str(hasta_id), f"Hasta ID: {hasta_id}")
                
                tedavi_hekimler = set(hasta_tedavi["DOKTOR_ID"].tolist())
                tahsilat_hekimler = set(hasta_tahsilat["DOKTOR_ID"].tolist())
                tum_hekimler = tedavi_hekimler.union(tahsilat_hekimler)
                
                for hekim_id in tum_hekimler:
                    hekim_tedavi = hasta_tedavi[hasta_tedavi["DOKTOR_ID"] == hekim_id]
                    tedavi_tutari = float(hekim_tedavi["TEDAVI_TOPLAM"].iloc[0]) if not hekim_tedavi.empty else 0.0
                    
                    hekim_tahsilat = hasta_tahsilat[hasta_tahsilat["DOKTOR_ID"] == hekim_id]
                    tahsilat_tutari = float(hekim_tahsilat["TOPLAM_TUTAR"].iloc[0]) if not hekim_tahsilat.empty else 0.0
                    son_tahsilat = str(hekim_tahsilat["SON_TAHSILAT_TARIHI"].iloc[0]) if not hekim_tahsilat.empty else "Yok"
                    
                    hekim_adi = hekim_tedavi["HEKIM_ADI"].iloc[0] if not hekim_tedavi.empty else f"Hekim ID: {hekim_id}"
                    
                    fark = tahsilat_tutari - tedavi_tutari
                    
                    if tedavi_tutari > 0 or tahsilat_tutari > 0:
                        ws2.cell(row=current_row, column=1, value=hasta_adi).border = border
                        ws2.cell(row=current_row, column=2, value=hasta_id).border = border
                        ws2.cell(row=current_row, column=3, value=hekim_adi).border = border
                        ws2.cell(row=current_row, column=4, value=hekim_id).border = border
                        ws2.cell(row=current_row, column=5, value=tedavi_tutari).number_format = '#,##0.00'
                        ws2.cell(row=current_row, column=5).border = border
                        ws2.cell(row=current_row, column=6, value=tahsilat_tutari).number_format = '#,##0.00'
                        ws2.cell(row=current_row, column=6).border = border
                        ws2.cell(row=current_row, column=7, value=fark).number_format = '#,##0.00'
                        ws2.cell(row=current_row, column=7).border = border
                        
                        # Fark negatifse kırmızı renk
                        if fark < 0:
                            ws2.cell(row=current_row, column=7).font = Font(color="FF0000")
                        elif fark > 0:
                            ws2.cell(row=current_row, column=7).font = Font(color="008000")
                        
                        ws2.cell(row=current_row, column=8, value=son_tahsilat).border = border
                        current_row += 1
            
            # Sütun genişliklerini ayarla
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 12
            ws2.column_dimensions['C'].width = 25
            ws2.column_dimensions['D'].width = 12
            ws2.column_dimensions['E'].width = 15
            ws2.column_dimensions['F'].width = 15
            ws2.column_dimensions['G'].width = 15
            ws2.column_dimensions['H'].width = 15
        
        # Excel dosyasını kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'tahsilat_raporu_{start_date}_{end_date}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Excel indirme hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": f"Excel oluşturulamadı: {str(e)}"}), 500  
@app.route("/api/personel/sil/<int:personel_id>", methods=["DELETE"])
@admin_required
def personel_sil_api(personel_id):
    try:
        data = request.get_json() or {}
        admin_sifre = data.get('admin_sifre', '').strip()
        
        if not personel_id:
            return jsonify({"error": "Personel ID gerekli"}), 400
            
        if not admin_sifre:
            return jsonify({"error": "Admin şifresi gerekli"}), 400
        
        # Admin şifresini kontrol et
        current_user = session.get("username")
        if current_user not in USERS:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 403
            
        if USERS[current_user]["password"] != admin_sifre:
            return jsonify({"error": "Hatalı admin şifresi"}), 403
        
        # Personeli sil
        success, message = personel_db.personel_sil(personel_id, admin_sifre)
        
        if success:
            return jsonify({"success": True, "message": message})
        else:
            return jsonify({"error": message}), 400
            
    except Exception as e:
        logger.error(f"Personel silme API hatası: {e}")
        return jsonify({"error": "Silme işlemi başarısız"}), 500
@app.route("/api/personel/maas/sil", methods=["POST"])
@admin_required
def personel_maas_sil():
    """Maaş kaydını ve ilişkili cari hareketini sil"""
    try:
        data = request.get_json() or {}
        
        maas_id = data.get('maas_id')
        admin_sifre = data.get('admin_sifre', '').strip()
        
        if not maas_id or not admin_sifre:
            return jsonify({"error": "Maaş ID ve admin şifresi gerekli"}), 400
        
        # Admin şifre kontrolü
        current_user = session.get("username")
        if current_user not in USERS or USERS[current_user]["password"] != admin_sifre:
            return jsonify({"error": "Hatalı admin şifresi"}), 403
        
        # Maaş kaydını getir
        import sqlite3
        conn = sqlite3.connect(personel_db.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT cari_id, personel_id, donem_ay, donem_yil FROM maas_odeme WHERE id = ?", (maas_id,))
        maas = cursor.fetchone()
        
        if not maas:
            conn.close()
            return jsonify({"error": "Maaş kaydı bulunamadı"}), 404
        
        cari_id, personel_id, donem_ay, donem_yil = maas
        
        # 1. İlişkili cari hareketini bul ve sil
        if cari_id:
            cursor.execute("""
                SELECT id FROM cari_hareketler 
                WHERE cari_id = ? 
                AND hareket_tipi = 'maas_alacak' 
                AND aciklama LIKE ?
            """, (cari_id, f"%{donem_ay}/{donem_yil}%"))
            
            hareket = cursor.fetchone()
            if hareket:
                # Cari hareketi sil ve bakiyeyi güncelle
                success, msg = cari_db.cari_hareket_sil(hareket[0], cari_id)
                if not success:
                    conn.close()
                    return jsonify({"error": f"Cari hareketi silinemedi: {msg}"}), 500
        
        # 2. Maaş kaydını sil
        cursor.execute("DELETE FROM maas_odeme WHERE id = ?", (maas_id,))
        conn.commit()
        conn.close()
        
        logger.info(f"Maaş kaydı ve cari hareketi silindi: Maaş ID {maas_id}")
        return jsonify({
            "success": True,
            "message": "Maaş kaydı ve ilişkili cari hareketi başarıyla silindi"
        })
        
    except Exception as e:
        logger.error(f"Maaş silme hatası: {e}")
        return jsonify({"error": f"Maaş silinirken hata oluştu: {str(e)}"}), 500
        
# --- TEDAVİLER ANALİZİ ---
@app.route("/analyze_tedaviler", methods=["POST"])
@login_required
def analyze_tedaviler():
    try:
        start_date_raw = request.form.get("start_date")
        end_date_raw = request.form.get("end_date")
        selected_branches = request.form.getlist("branches")
        selected_doctors = request.form.getlist("doctors")
        
        if not start_date_raw or not end_date_raw:
            flash("Başlangıç ve bitiş tarihi gereklidir.", "danger")
            return redirect(url_for("tedaviler"))

        try:
            start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_raw, "%Y-%m-%d").date()
        except ValueError:
            flash("Tarih formatı hatalı.", "danger")
            return redirect(url_for("tedaviler"))

        engine = get_database_connection()

        params = {
            'start_date': start_date,
            'end_date': end_date,
            'silindi': 'False',
            'isdeleted': 'False'
        }

        # DOSYA_NO KALDIRILDI
        base_query = """
        SELECT
            T.ROWNO,
            T.HASTA_ID,
            CONCAT(IFNULL(H.ADI,''), ' ', IFNULL(H.SOYADI,'')) AS HASTA_ADI,
            T.DOKTOR_ID,
            CONCAT(IFNULL(DR.ADI,''), ' ', IFNULL(DR.SOYADI,'')) AS HEKIM_ADI,
            T.SUBE_ID,
            IFNULL(S.UNVANI, 'Bilinmeyen') AS SUBE_ADI,
            T.TARIH,
            T.ISLEM,
            T.TEDAVIADI,
            IFNULL(T.TUTAR, 0) AS TUTAR,
            IFNULL(T.LISTETUTAR, 0) AS LISTETUTAR,
            T.PARABIRIMI
        FROM tedavi AS T
        LEFT JOIN karthasta AS H ON T.HASTA_ID = H.HASTA_ID
        LEFT JOIN kartdoktor AS DR ON T.DOKTOR_ID = DR.CARI_ID
        LEFT JOIN subeler AS S ON T.SUBE_ID = S.CARI_ID
        WHERE T.SILINDI = :silindi
          AND T.ISDELETED = :isdeleted
          AND T.TARIH BETWEEN :start_date AND :end_date
        """

        if selected_branches:
            branch_placeholders = ','.join([f':branch_{i}' for i in range(len(selected_branches))])
            base_query += f" AND T.SUBE_ID IN ({branch_placeholders})"
            for i, branch in enumerate(selected_branches):
                params[f'branch_{i}'] = branch

        if selected_doctors:
            doctor_placeholders = ','.join([f':doctor_{i}' for i in range(len(selected_doctors))])
            base_query += f" AND T.DOKTOR_ID IN ({doctor_placeholders})"
            for i, doctor in enumerate(selected_doctors):
                params[f'doctor_{i}'] = doctor

        base_query += " ORDER BY T.TARIH, T.SUBE_ID, T.DOKTOR_ID"

        df = pd.read_sql(text(base_query), engine, params=params)

        if df.empty:
            branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
            branches_df = pd.read_sql(branches_query, engine)
            branches = branches_df.to_dict('records')
            
            return render_template("tedaviler.html",
                                 branches=branches,
                                 error_date="Bu kriterlere uygun tedavi kaydı bulunamadı.")

        ozet_df = df.groupby(['HEKIM_ADI', 'SUBE_ADI', 'ISLEM', 'TEDAVIADI']).agg({
            'TUTAR': 'sum',
            'LISTETUTAR': 'sum',
            'ROWNO': 'count'
        }).reset_index()
        ozet_df.columns = ['HEKIM_ADI', 'SUBE_ADI', 'ISLEM', 'TEDAVIADI', 'TOPLAM_TUTAR', 'TOPLAM_LISTE', 'ADET']

        # DOSYA_NO KALDIRILDI
        hasta_detay = df[['HASTA_ADI', 'HEKIM_ADI', 'SUBE_ADI', 'TARIH', 
                         'TEDAVIADI', 'TUTAR', 'LISTETUTAR']].copy()

        toplam_tedavi_sayisi = len(df)
        toplam_tutar = df['TUTAR'].sum()
        toplam_liste = df['LISTETUTAR'].sum()
        toplam_indirim = toplam_liste - toplam_tutar

        branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
        branches_df = pd.read_sql(branches_query, engine)
        branches = branches_df.to_dict('records')

        return render_template("tedaviler.html",
                             branches=branches,
                             ozet_data=ozet_df.to_dict('records'),
                             hasta_data=hasta_detay.to_dict('records'),
                             toplam_tedavi=toplam_tedavi_sayisi,
                             toplam_tutar=f"{toplam_tutar:,.2f}",
                             toplam_liste=f"{toplam_liste:,.2f}",
                             toplam_indirim=f"{toplam_indirim:,.2f}",
                             start_date=start_date,
                             end_date=end_date,
                             selected_branches=selected_branches,
                             selected_doctors=selected_doctors)

    except Exception as e:
        logger.error(f"Tedavi analizi hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        
        engine = get_database_connection()
        branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
        branches_df = pd.read_sql(branches_query, engine)
        branches = branches_df.to_dict('records')
        
        flash("Analiz sırasında bir hata oluştu.", "danger")
        return render_template("tedaviler.html", 
                             branches=branches,
                             error_date=f"Hata: {str(e)}")
                             
@app.route("/download_tedaviler", methods=["POST"])
@login_required
def download_tedaviler():
    """Tedavi analizini Excel olarak indir"""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        data = request.get_json() or {}
        
        ozet_data = data.get("ozet_data", [])
        hasta_data = data.get("hasta_data", [])
        
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        
        # Border stili
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Sayfa: Tedavi Türü Özeti
        ws1 = wb.active
        ws1.title = "Tedavi Türü Özeti"
        
        # Başlıklar
        headers_ozet = ["HEKİM ADI", "ŞUBE ADI", "TEDAVİ TÜRÜ", "TEDAVİ ADI", 
                       "ADET", "TOPLAM TUTAR", "LİSTE TUTARI", "İNDİRİM"]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers_ozet, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Veri satırları
        for row_num, row_data in enumerate(ozet_data, 2):
            indirim = float(row_data['TOPLAM_LISTE']) - float(row_data['TOPLAM_TUTAR'])
            
            ws1.cell(row=row_num, column=1, value=row_data['HEKIM_ADI']).border = border
            ws1.cell(row=row_num, column=2, value=row_data['SUBE_ADI']).border = border
            ws1.cell(row=row_num, column=3, value=row_data['ISLEM']).border = border
            ws1.cell(row=row_num, column=4, value=row_data['TEDAVIADI']).border = border
            ws1.cell(row=row_num, column=5, value=int(row_data['ADET'])).border = border
            ws1.cell(row=row_num, column=6, value=float(row_data['TOPLAM_TUTAR'])).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=6).border = border
            ws1.cell(row=row_num, column=7, value=float(row_data['TOPLAM_LISTE'])).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=7).border = border
            ws1.cell(row=row_num, column=8, value=indirim).number_format = '#,##0.00'
            ws1.cell(row=row_num, column=8).border = border
        
        # Sütun genişlikleri
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 35
        ws1.column_dimensions['E'].width = 10
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        
        # 2. Sayfa: Hasta Detayları
        ws2 = wb.create_sheet(title="Hasta Detayları")
        
        headers_hasta = ["DOSYA NO", "HASTA ADI", "HEKİM ADI", "ŞUBE ADI", 
                        "TARİH", "TEDAVİ ADI", "TUTAR", "LİSTE FİYATI"]
        
        for col_num, header in enumerate(headers_hasta, 1):
            cell = ws2.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Veri satırları
        for row_num, row_data in enumerate(hasta_data, 2):
            ws2.cell(row=row_num, column=1, value=row_data.get('DOSYA_NO', '')).border = border
            ws2.cell(row=row_num, column=2, value=row_data['HASTA_ADI']).border = border
            ws2.cell(row=row_num, column=3, value=row_data['HEKIM_ADI']).border = border
            ws2.cell(row=row_num, column=4, value=row_data['SUBE_ADI']).border = border
            ws2.cell(row=row_num, column=5, value=str(row_data['TARIH'])).border = border
            ws2.cell(row=row_num, column=6, value=row_data['TEDAVIADI']).border = border
            ws2.cell(row=row_num, column=7, value=float(row_data['TUTAR'])).number_format = '#,##0.00'
            ws2.cell(row=row_num, column=7).border = border
            ws2.cell(row=row_num, column=8, value=float(row_data['LISTETUTAR'])).number_format = '#,##0.00'
            ws2.cell(row=row_num, column=8).border = border
        
        # Sütun genişlikleri
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 20
        ws2.column_dimensions['E'].width = 12
        ws2.column_dimensions['F'].width = 35
        ws2.column_dimensions['G'].width = 15
        ws2.column_dimensions['H'].width = 15
        
        # Excel dosyasını kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        start_date = data.get('start_date', 'baslangic')
        end_date = data.get('end_date', 'bitis')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'tedavi_raporu_{start_date}_{end_date}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Excel indirme hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": f"Excel oluşturulamadı: {str(e)}"}), 500
# --- TEDAVİLER SAYFASI ---
@app.route("/tedaviler")
@login_required
def tedaviler():
    """Tedaviler analizi ana sayfası"""
    try:
        engine = get_database_connection()
        
        # Şubeleri getir
        branches_query = "SELECT CARI_ID, UNVANI FROM subeler WHERE SILINDI='false' ORDER BY UNVANI"
        branches_df = pd.read_sql(branches_query, engine)
        branches = branches_df.to_dict('records')
        
        return render_template("tedaviler.html",
                             branches=branches,
                             ozet_data=None,
                             hasta_data=None,
                             toplam_tedavi=None,
                             toplam_tutar=None,
                             toplam_liste=None,
                             toplam_indirim=None,
                             start_date=None,
                             end_date=None,
                             selected_branches=None,
                             selected_doctors=None,
                             error_date=None)
    except Exception as e:
        logger.error(f"Tedaviler sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu.", "danger")
        return redirect(url_for("home"))

# ==================== İZİN YÖNETİMİ ROUTES ====================

@app.route("/izin_yonetimi")
@login_required
def izin_yonetimi():
    """İzin yönetimi ana sayfası"""
    if session.get("role") != "admin":
        flash("Bu sayfaya erişim yetkiniz yok", "danger")
        return redirect(url_for("home"))
    
    try:
        # Şubeleri getir
        query = "SELECT CARI_ID AS id, UNVANI AS name FROM subeler WHERE SILINDI = :silindi ORDER BY UNVANI"
        df = execute_query(query, {"silindi": "false"})
        branches = df.to_dict("records")
        
        # Aktif personelleri getir
        personeller = personel_db.personel_listele(sadece_aktif=True)
        
        return render_template("izin_yonetimi.html", 
                             branches=branches,
                             personeller=personeller)
        
    except Exception as e:
        logger.error(f"İzin yönetimi sayfası yüklenirken hata: {e}")
        flash("Sayfa yüklenirken bir hata oluştu", "danger")
        return redirect(url_for("home"))

@app.route("/api/izin/ekle", methods=["POST"])
@admin_required
def izin_ekle_api():
    """Yeni izin kaydı ekle"""
    try:
        data = request.get_json() or {}
        
        # Gerekli alanları kontrol et
        required_fields = ["personel_id", "izin_tipi", "baslangic_tarihi", "bitis_tarihi", "gun_sayisi"]
        valid, error_msg = validate_required_fields(data, required_fields)
        if not valid:
            return jsonify({"error": error_msg}), 400
        
        # İzin ekle
        success, result = personel_db.izin_ekle(data)
        
        if success:
            return jsonify({
                "success": True,
                "message": "İzin başarıyla kaydedildi",
                "izin_id": result
            })
        else:
            return jsonify({"error": result}), 400
        
    except Exception as e:
        logger.error(f"İzin ekleme API hatası: {e}")
        return jsonify({"error": "İzin eklenirken hata oluştu"}), 500

@app.route("/api/izin/liste")
@login_required
def izin_liste_api():
    """İzinleri listele"""
    try:
        personel_id = request.args.get("personel_id")
        baslangic = request.args.get("baslangic")
        bitis = request.args.get("bitis")
        izin_tipi = request.args.get("izin_tipi")
        
        izinler = personel_db.izin_listele(
            personel_id=personel_id,
            baslangic=baslangic,
            bitis=bitis,
            izin_tipi=izin_tipi
        )
        
        return jsonify({
            "success": True,
            "data": izinler
        })
        
    except Exception as e:
        logger.error(f"İzin listeleme API hatası: {e}")
        return jsonify({"error": "İzinler yüklenemedi"}), 500

@app.route("/api/izin/sil/<int:izin_id>", methods=["DELETE"])
@admin_required
def izin_sil_api(izin_id):
    """İzin kaydını sil"""
    try:
        success, message = personel_db.izin_sil(izin_id)
        
        if success:
            return jsonify({
                "success": True,
                "message": message
            })
        else:
            return jsonify({"error": message}), 400
        
    except Exception as e:
        logger.error(f"İzin silme API hatası: {e}")
        return jsonify({"error": "İzin silinemedi"}), 500

@app.route("/api/izin/yillik_durum/<int:personel_id>")
@login_required
def yillik_izin_durum_api(personel_id):
    """Personelin yıllık izin durumunu getir"""
    try:
        durum = personel_db.yillik_izin_durumu(personel_id)
        
        if durum:
            return jsonify({
                "success": True,
                "data": durum
            })
        else:
            return jsonify({
                "success": False,
                "error": "Personel bulunamadı"
            }), 404
        
    except Exception as e:
        logger.error(f"Yıllık izin durum API hatası: {e}")
        return jsonify({"error": "Yıllık izin durumu alınamadı"}), 500

@app.route("/api/izin/aylik_kesinti", methods=["POST"])
@login_required
def aylik_kesinti_api():
    """Belirli ay için kesintili izin günlerini hesapla"""
    try:
        data = request.get_json() or {}
        
        personel_id = data.get("personel_id")
        donem_ay = data.get("donem_ay")
        donem_yil = data.get("donem_yil")
        
        if not all([personel_id, donem_ay, donem_yil]):
            return jsonify({"error": "Eksik parametreler"}), 400
        
        kesintili_gun = personel_db.aylik_kesintili_izin_hesapla(
            personel_id,
            donem_ay,
            donem_yil
        )
        
        return jsonify({
            "success": True,
            "kesintili_gun": kesintili_gun
        })
        
    except Exception as e:
        logger.error(f"Aylık kesinti API hatası: {e}")
        return jsonify({"error": "Kesinti hesaplanamadı"}), 500

@app.route("/api/izin/ozet_rapor")
@login_required
def izin_ozet_rapor_api():
    """İzin özet raporu"""
    try:
        personel_id = request.args.get("personel_id")
        donem_yil = request.args.get("donem_yil")
        
        rapor = personel_db.izin_ozet_rapor(
            personel_id=personel_id,
            donem_yil=donem_yil
        )
        
        return jsonify({
            "success": True,
            "data": rapor
        })
        
    except Exception as e:
        logger.error(f"İzin rapor API hatası: {e}")
        return jsonify({"error": "Rapor oluşturulamadı"}), 500

@app.route("/api/izin/excel_export")
@login_required
def izin_excel_export():
    """İzin raporunu Excel olarak indir"""
    try:
        import io
        from datetime import datetime
        
        donem_yil = request.args.get("donem_yil", datetime.now().year)
        
        rapor = personel_db.izin_ozet_rapor(donem_yil=donem_yil)
        
        if not rapor:
            flash("Export için veri bulunamadı", "warning")
            return redirect(url_for("izin_yonetimi"))
        
        # DataFrame oluştur
        df = pd.DataFrame(rapor)
        
        # Kolon isimleri düzenle
        df = df.rename(columns={
            'ad': 'Ad',
            'soyad': 'Soyad',
            'yillik_toplam_hak': 'Yıllık Hak',
            'yillik_kullanilan': 'Kullanılan',
            'yillik_kalan': 'Kalan',
            'ucretsiz_gun': 'Ücretsiz İzin',
            'hastalik_gun': 'Hastalık İzni',
            'mazeret_gun': 'Mazeret İzni',
            'toplam_kesintili': 'Kesintili Gün'
        })
        
        # Excel dosyası oluştur
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='İzin Raporu', index=False)
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'izin_raporu_{donem_yil}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Excel export hatası: {e}")
        flash("Excel dosyası oluşturulamadı", "danger")
        return redirect(url_for("izin_yonetimi"))        

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)        

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)